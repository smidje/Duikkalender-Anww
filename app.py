# app.py — ANWW Duikapp (login via username + wachtwoord) • Build v2025-10-20
# Functies: Activiteitenkalender (met inschrijven + tot 3 maaltijdkeuzes + opmerking),
# Afrekening (alleen duikers tellen mee), Duiken invoeren/overzicht, Beheer (leden),
# Wekelijkse mail preview/export (eerstvolgende 4 activiteiten).
#
# Minimale, gerichte aanpassingen in deze versie:
# - BREVET_CHOICES exact: ['k1ster','1ster','2ster','3ster','4ster','ass-inst','1*inst','2*inst','3*inst']
# - canon_brevet/normalize_brevet mappen varianten naar deze 9 waarden
# - Datumselectie bij activiteit: dag/maand/jaar (DD/MM/YYYY)
# - Opmerkingenveld bij activiteit + mee in export/afdruk
# - Duiken: overnemen van datum/plaats/duikers uit activiteit (indien beschikbaar)
# - Verwijderd: dubbele st.date_input('Datum*') om StreamlitDuplicateElementId te vermijden

import streamlit as st
import requests
import pandas as pd
import datetime
from datetime import datetime as dt
from openpyxl.styles import Font, Alignment
import io
import os
import time
import math
import hmac, hashlib, secrets
from typing import Optional
from supabase import create_client, Client
from postgrest.exceptions import APIError
import httpx
import json
import base64
# --- Banner helpers (RPC eerst; REST/DB/sessie fallback) ---
import base64, os, json
from datetime import datetime as _dt

def _sb_rest_conf():
    url = os.getenv('SUPABASE_URL')
    sk = os.getenv('SUPABASE_SERVICE_KEY')
    if url and sk:
        return url.rstrip('/'), sk
    return None, None

def _asset_rest_get(key: str):
    url, sk = _sb_rest_conf()
    if not (url and sk):
        return None
    try:
        r = httpx.get(
            f'{url}/rest/v1/app_assets',
            params={'key': f'eq.{key}', 'select': 'image_b64', 'limit': '1'},
            headers={'apikey': sk, 'Authorization': f'Bearer {sk}'},
            timeout=10.0,
        )
        if r.status_code < 400:
            rows = r.json() or []
            return (rows[0].get('image_b64') if rows else None)
    except Exception:
        pass
    return None

def _asset_rest_upsert(key: str, value_text: str):
    url, sk = _sb_rest_conf()
    if not (url and sk):
        return False
    payload = {'key': key, 'image_b64': value_text, 'updated_by': None, 'updated_ts': _dt.utcnow().isoformat()}
    try:
        r = httpx.post(
            f'{url}/rest/v1/app_assets',
            json=payload,
            params={'on_conflict': 'key'},
            headers={
                'apikey': sk,
                'Authorization': f'Bearer {sk}',
                'Content-Type': 'application/json',
                'Prefer': 'resolution=merge-duplicates',
            },
            timeout=10.0,
        )
        return r.status_code < 400
    except Exception:
        return False

def _asset_rest_delete(key: str):
    url, sk = _sb_rest_conf()
    if not (url and sk):
        return False
    try:
        r = httpx.delete(
            f'{url}/rest/v1/app_assets',
            params={'key': f'eq.{key}'},
            headers={'apikey': sk, 'Authorization': f'Bearer {sk}'},
            timeout=10.0,
        )
        return r.status_code < 400
    except Exception:
        return False

def asset_get_bytes(key: str) -> bytes | None:
    txt = _asset_rest_get(key)
    if txt:
        try:
            return base64.b64decode(txt)
        except Exception:
            pass
    try:
        res = run_db(lambda c: c.table('app_assets').select('image_b64').eq('key', key).limit(1).execute(), what=f'app_assets get {key}')
        rows = getattr(res, 'data', None) or []
        if rows and rows[0].get('image_b64'):
            return base64.b64decode(rows[0]['image_b64'])
    except Exception:
        pass
    return st.session_state.get(f'_asset_{key}')

def asset_set_bytes(key: str, data: bytes | None, updated_by: str | None = None):
    if is_readonly():
        raise Exception('Read-only modus')
    try:
        if data is None:
            run_db(lambda c: c.rpc('app_assets_delete', {'k': key}).execute(), what=f'rpc app_assets_delete {key}')
            return
        b64 = base64.b64encode(data).decode('ascii')
        run_db(lambda c: c.rpc('app_assets_upsert', {'k': key, 'v': b64, 'who': (updated_by or None)}).execute(), what=f'rpc app_assets_upsert {key}')
        return
    except Exception:
        pass
    if data is None:
        if _asset_rest_delete(key):
            return
    else:
        b64 = base64.b64encode(data).decode('ascii')
        if _asset_rest_upsert(key, b64):
            return
    try:
        if data is None:
            run_db(lambda c: c.table('app_assets').delete().eq('key', key).execute(), what=f'app_assets delete {key}')
        else:
            b64 = base64.b64encode(data).decode('ascii')
            run_db(lambda c: c.table('app_assets').upsert({'key': key, 'image_b64': b64}, on_conflict='key').execute(), what=f'app_assets upsert {key}')
    except Exception:
        st.session_state[f'_asset_{key}'] = (None if data is None else data)

def asset_meta_get(bkey: str) -> dict:
    mkey = f'{bkey}_meta'
    txt = _asset_rest_get(mkey)
    if not txt:
        try:
            res = run_db(lambda c: c.table('app_assets').select('image_b64').eq('key', mkey).limit(1).execute(), what=f'app_assets get {mkey}')
            rows = getattr(res, 'data', None) or []
            txt = rows[0]['image_b64'] if (rows and rows[0].get('image_b64')) else None
        except Exception:
            txt = None
    if txt:
        try:
            raw = base64.b64decode(txt).decode('utf-8')
            return json.loads(raw)
        except Exception:
            return {}
    return st.session_state.get(f'_asset_{mkey}', {})

def asset_meta_set(bkey: str, meta: dict):
    if is_readonly():
        raise Exception('Read-only modus')
    mkey = f'{bkey}_meta'
    try:
        raw = json.dumps(meta)
        b64 = base64.b64encode(raw.encode('utf-8')).decode('ascii')
    except Exception:
        b64 = base64.b64encode(b'{}').decode('ascii')
    try:
        run_db(lambda c: c.rpc('app_assets_upsert', {'k': mkey, 'v': b64, 'who': None}).execute(), what=f'rpc app_assets_upsert {mkey}')
        return
    except Exception:
        pass
    if _asset_rest_upsert(mkey, b64):
        return
    try:
        run_db(lambda c: c.table('app_assets').upsert({'key': mkey, 'image_b64': b64}, on_conflict='key').execute(), what=f'app_assets upsert {mkey}')
    except Exception:
        st.session_state[f'_asset_{mkey}'] = meta

def _banner_html(img_bytes: bytes, meta: dict) -> str:
    if not img_bytes:
        return ""
    b64 = base64.b64encode(img_bytes).decode('ascii')
    link = (meta.get('link_url') or "").strip()
    width_mode = meta.get('width_mode') or 'full'
    width_px = int(meta.get('width_px') or 0)
    height_px = int(meta.get('height_px') or 0)
    align = (meta.get('align') or 'center').lower()

    css_w = f"{width_px}px" if (width_mode == 'px' and width_px > 0) else "100%"
    css_h = f"{height_px}px" if height_px > 0 else "auto"
    if align == 'left':
        css_margin = "0 auto 0 0"
    elif align == 'right':
        css_margin = "0 0 0 auto"
    else:
        css_margin = "0 auto"

    img_tag = f"<img src='data:image/jpeg;base64,{b64}' style='display:block;width:{css_w};height:{css_h};object-fit:contain;margin:{css_margin};' />"
    if link:
        return f"<a href='{link}' target='_blank' rel='noopener'>{img_tag}</a>"
    return img_tag


st.set_page_config(page_title="ANWW Duikapp", layout="wide")
APP_BUILD = "v2025-10-20"
APP_PUBLIC_URL = os.getenv("APP_PUBLIC_URL", "").rstrip("/")

def inject_css():
    st.markdown("""
    <style>
      :root { --background: #8DAEBA; --secondary: #A38B16; --text: #11064D; --primary: #728DCC; --border: #2a355a; --success: #3CA133; --warning: #f59e0b; --error: #ef4444; }
      .stApp, [data-testid="stAppViewContainer"], section.main, div.block-container { background-color: var(--background) !important; color: var(--text) !important; }
      section[data-testid="stSidebar"], [data-testid="stSidebarContent"] { background-color: var(--secondary) !important; }
      .stTabs [data-baseweb="tab"] { background: var(--secondary) !important; color: #fff !important; border-radius: 5px 5px 0 0; font-weight: 600; }
      .stTabs [aria-selected="true"] { background: var(--primary) !important; color: white !important; }
      .stExpander { border: 1px solid var(--border) !important; background: #ffffff80 !important; }
      .stButton > button, .stDownloadButton > button { background-color: var(--primary) !important; color: #fff !important; border: 2px solid var(--border) !important; border-radius: 10px !important; [...]
      .stButton > button:hover, .stDownloadButton > button:hover { filter: brightness(1.1) !important; transform: translateY(-1px); }
      .stButton.success > button { background: var(--success) !important; border-color: var(--success) !important; }
      .stButton.warning > button { background: var(--warning) !important; color:#000 !important; border-color: var(--warning) !important; }
      .stButton.error   > button { background: var(--error)   !important; border-color: var(--error)   !important; }
      [data-testid="stDataFrame"] thead tr th { background: #ffffff88 !important; color: var(--text) !important; }
      hr, .stDivider { border-top: 2px solid var(--border) !important; }
      h1, h2, h3, h4, h5 { color: var(--text) !important; }
    </style>
    """, unsafe_allow_html=True)
inject_css()

def is_readonly() -> bool:
    return bool(st.secrets.get("app", {}).get("force_readonly", False))

@st.cache_resource
def get_client() -> Client:
    supa = st.secrets.get("supabase", {})
    url = supa.get("url") or os.getenv("SUPABASE_URL")
    key = supa.get("anon_key") or os.getenv("SUPABASE_ANON_KEY")
    if not url or not key:
        st.error("Supabase credentials ontbreken. Zet [supabase].url en anon_key in secrets.toml.")
        st.stop()
    return create_client(url, key)

sb: Client = get_client()

def run_db(fn, *, what="db call", tries=2, backoff=0.4):
    for i in range(tries):
        try:
            return fn(sb)
        except (httpx.ConnectError, httpx.ReadTimeout) as e:
            if i+1 < tries:
                time.sleep(backoff * (i+1))
                continue
            st.error(f"Netwerkfout bij {what}: {e}"); st.stop()
        except APIError as e:
            st.error(f"API-fout bij {what}. Controleer tabellen/RLS/policies (anon).");
            st.caption(str(e)); st.stop()
        except Exception as e:
            if i+1 < tries:
                time.sleep(backoff * (i+1))
                continue
            st.error(f"Onverwachte fout bij {what}: {e}"); st.stop()

def _hash_password(password: str, iterations: int = 240000) -> str:
    salt = secrets.token_bytes(16)
    dk = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, iterations)
    return f"pbkdf2${iterations}${salt.hex()}${dk.hex()}"

def _verify_password(password: str, hashed: str) -> bool:
    try:
        algo, iters, salt_hex, hash_hex = hashed.split("$", 3)
        if algo != "pbkdf2":
            return False
        iters = int(iters)
        salt = bytes.fromhex(salt_hex)
        ref = bytes.fromhex(hash_hex)
        test = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, iters)
        return hmac.compare_digest(ref, test)
    except Exception:
        return False

def auth_get_user_by_username(login: str) -> Optional[dict]:
    login = (login or "").strip()
    if not login:
        return None
    res = run_db(lambda c: c.table("auth_local").select("*").eq("username", login).limit(1).execute(),
                 what="auth_local select by username")
    rows = res.data or []
    if rows:
        return rows[0]
    if "@" in login:
        res2 = run_db(lambda c: c.table("auth_local").select("*").eq("email", login.lower()).limit(1).execute(),
                      what="auth_local select by email")
        rows2 = res2.data or []
        if rows2:
            return rows2[0]
    return None

def auth_count_admins() -> int:
    res = run_db(lambda c: c.table("auth_local").select("id", count="exact").eq("role","admin").execute(),
                 what="auth_local count admin")
    return res.count or 0

def auth_create_user(username: str, password: str, role: str, email: str | None = None):
    if is_readonly():
        raise Exception("Read-only modus")
    role = (role or "viewer").strip().lower()
    if role not in {"admin","user","member","viewer"}:
        role = "viewer"
    payload = {
        "username": (username or "").strip(),
        "email": (email or "").strip().lower() or None,
        "password_hash": _hash_password(password),
        "role": role,
    }
    run_db(lambda c: c.table("auth_local").insert(payload).execute(), what="auth_local insert")

def auth_update_password_by_username(username: str, new_pw: str):
    if is_readonly():
        raise Exception("Read-only modus")
    run_db(lambda c: c.table("auth_local").update({"password_hash": _hash_password(new_pw)}).eq("username", username).execute(),
           what="auth_local update password")

def current_user():
    return st.session_state.get("auth_user") or {}

def current_role() -> str:
    u = current_user()
    r = (u.get("role") or "viewer").lower()
    return r if r in {"admin","user","member","viewer"} else "viewer"

def current_username() -> str:
    return (current_user().get("username") or "").strip()

def current_email() -> str:
    return (current_user().get("email") or "").strip().lower()

BREVET_CHOICES = ['k1ster','1ster','2ster','3ster','4ster','ass-inst','1*inst','2*inst','3*inst']

def normalize_brevet(v: str | None) -> str | None:
    if v is None:
        return None
    s = str(v).strip().lower()
    if s in {"", "(geen)", "geen", "none", "null"}:
        return None
    import re as _re
    s = _re.sub(r'[-–—‒−]', '-', s)
    s = _re.sub(r'[＊∗⋆✱✳︎]', '*', s)
    s = s.replace(" ", "")
    s = s.replace("instructeur","inst").replace("assistent","ass")
    if s in {"k1","k1ster","kand1","kand1ster","kandidaat1ster"}: return "k1ster"
    if s in {"1","1ster","1*"}: return "1ster"
    if s in {"2","2ster","2*"}: return "2ster"
    if s in {"3","3ster","3*"}: return "3ster"
    if s in {"4","4ster","4*"}: return "4ster"
    if s in {"assinst","ass-inst","assinst.","assinstr"}: return "ass-inst"
    if s in {"1inst","1*inst","1-inst","1instructeur","1*instructeur"}: return "1*inst"
    if s in {"2inst","2*inst","2-inst","2instructeur","2*instructeur"}: return "2*inst"
    if s in {"3inst","3*inst","3-inst","3instructeur","3*instructeur"}: return "3*inst"
    return s

def canon_brevet(v):
    if v is None:
        return None
    s = normalize_brevet(v)
    if s is None:
        return None
    xmap = {"assinst":"ass-inst","1inst":"1*inst","2inst":"2*inst","3inst":"3*inst","k1":"k1ster"}
    s = xmap.get(s, s)
    return s if s in BREVET_CHOICES else None

ROLE_CHOICES = ['admin','user','member','viewer']


def leden_upsert(payload: dict):
    if is_readonly():
        raise Exception("Read-only modus")
    p = dict(payload)
    if "duikbrevet" in p:
        val = p.get("duikbrevet")
        p["duikbrevet"] = None if val is None else canon_brevet(val)
    if "email" in p:
        em = (p.get("email") or "").strip()
        p["email"] = em.lower() if em else None

    # NOTE:
    # Postgres upsert requires that 'on_conflict' matches an existing UNIQUE/PK constraint.
    # Many schemas have *separate* unique constraints (e.g. UNIQUE(email), UNIQUE(username)),
    # so 'email,username' is invalid unless there is a composite unique index.
    # We avoid the ON CONFLICT error by doing a small merge routine:
    # 1) If 'id' is present, upsert by primary key 'id'.
    # 2) Else, try to find an existing row by username, then by email; if found, update by id.
    # 3) If nothing exists, insert a new row.
    def _select_id_by(field: str, value: str):
        if not value:
            return None
        res = run_db(lambda c: c.table("leden").select("id").eq(field, value).limit(1).execute(), what=f"leden select by {field}")
        rows = (res.data or []) if hasattr(res, "data") else []
        return rows[0].get("id") if rows else None

    # Case 1: Upsert by primary key if we have it
    if p.get("id"):
        run_db(lambda c: c.table("leden").upsert(p, on_conflict="id").execute(), what="leden upsert by id")
        return

    # Case 2: Try to match existing by username/email and update
    uid = None
    if p.get("username"):
        uid = _select_id_by("username", p.get("username"))
    if not uid and p.get("email"):
        uid = _select_id_by("email", p.get("email"))

    if uid:
        run_db(lambda c: c.table("leden").update(p).eq("id", uid).execute(), what="leden update by id (merge)")
    else:
        run_db(lambda c: c.table("leden").insert(p).execute(), what="leden insert")

def leden_list_df() -> pd.DataFrame:
    res = run_db(lambda c: c.table("leden").select("*").order("achternaam").order("voornaam").execute(),
                 what="leden select")
    return pd.DataFrame(res.data or [])

def leden_get_by_email(email: str) -> Optional[dict]:
    email = (email or "").lower().strip()
    if not email:
        return None
    res = run_db(lambda c: c.table("leden").select("*").eq("email", email).limit(1).execute(),
                 what="leden by email")
    rows = res.data or []
    return rows[0] if rows else None

def leden_get_by_username(username: str) -> Optional[dict]:
    if not username:
        return None
    res = run_db(lambda c: c.table("leden").select("*").eq("username", username).limit(1).execute(),
                 what="leden by username")
    rows = res.data or []
    return rows[0] if rows else None

def duikers_labels() -> list[str]:
    res = run_db(lambda c: c.table("duikers").select("voornaam, achternaam, naam").execute(),
                 what="duikers select")
    rows = res.data or []
    out = []
    for r in rows:
        vn, an = (r.get("voornaam") or "").strip(), (r.get("achternaam") or "").strip()
        out.append(f"{an}, {vn}".strip(", ") if (vn or an) else (r.get("naam") or "").strip())
    def key(x):
        if "," in x:
            an, vn = [p.strip() for p in x.split(",", 1)]
            return (an.lower(), vn.lower())
        parts = x.split()
        return (parts[-1].lower() if parts else "", " ".join(parts[:-1]).lower())
    return sorted([o for o in out if o], key=key)

def plaatsen_list() -> list[str]:
    res = run_db(lambda c: c.table("duikplaatsen").select("plaats").order("plaats").execute(),
                 what="duikplaatsen select")
    return [r["plaats"] for r in (res.data or [])]

def plaats_add(plaats: str):
    if is_readonly():
        raise Exception("Read-only modus")
    run_db(lambda c: c.table("duikplaatsen").insert({"plaats": plaats}).execute(),
           what="duikplaatsen insert")

def duiken_insert(rows: list[dict]):
    if is_readonly():
        raise Exception("Read-only modus")
    if rows:
        run_db(lambda c: c.table("duiken").insert(rows).execute(), what="duiken insert")

def duiken_fetch_df() -> pd.DataFrame:
    res = run_db(lambda c: c.table("duiken").select("*").order("datum", desc=True).order("plaats").order("duiker").execute(),
                 what="duiken select")
    return pd.DataFrame(res.data or [])

def duiken_delete_by_ids(ids: list):
    if is_readonly():
        raise Exception("Read-only modus")
    if ids:
        run_db(lambda c: c.table("duiken").delete().in_("id", ids).execute(), what="duiken delete")

def afrekening_insert(row: dict):
    if is_readonly():
        raise Exception("Read-only modus")
    run_db(lambda c: c.table("afrekeningen").insert(row).execute(), what="afrekeningen insert")

def activiteit_add(titel, omschr, datum, tijd, locatie, meal_opts, created_by, opmerkingen=None, organisator=None):
    if is_readonly():
        raise Exception("Read-only modus")
    payload = {
        "titel": titel.strip(),
        "omschrijving": (omschr or "").strip(),
        "datum": datum.isoformat(),
        "tijd": tijd.isoformat() if tijd else None,
        "locatie": (locatie or "").strip() or None,
        "meal_options": meal_opts or None,
        "created_by": created_by or None,
        "opmerkingen": (opmerkingen or "").strip() or None,
        "organisator": (organisator or None)
    }
    run_db(lambda c: c.table("activiteiten").insert(payload).execute(), what="activiteiten insert")

def activiteit_update(activiteit_id: str, payload: dict):
    if is_readonly():
        raise Exception("Read-only modus")
    run_db(lambda c: c.table("activiteiten").update(payload).eq("id", activiteit_id).execute(),
           what="activiteiten update")

def signups_delete_users(activiteit_id: str, usernames: list[str] | None = None, lid_ids: list[str] | None = None):
    if is_readonly():
        raise Exception("Read-only modus")
    def _go(c):
        q = c.table("activity_signups").delete().eq("activiteit_id", activiteit_id)
        if usernames:
            q = q.in_("username", usernames)
        if lid_ids:
            q = q.in_("lid_id", lid_ids)
        return q.execute()
    return _go


def activiteiten_list_df(upcoming=True) -> pd.DataFrame:
    def _go(c):
        q = c.table("activiteiten").select("*")
        if upcoming:
            q = q.gte("datum", datetime.date.today().isoformat())
        return q.order("datum").order("tijd").execute()
    res = run_db(_go, what="activiteiten select")
    return pd.DataFrame(res.data or [])

def signups_get(activiteit_id: str) -> pd.DataFrame:
    res = run_db(lambda c: c.table("activity_signups").select("*").eq("activiteit_id", activiteit_id).order("signup_ts").execute(),
                 what="signups select")
    df = pd.DataFrame(res.data or [])
    for col in ["id","activiteit_id","username","lid_id","status","eating","meal_choice","opmerking","signup_ts"]:
        if col not in df.columns:
            df[col] = None
    try:
        df["signup_ts"] = pd.to_datetime(df["signup_ts"], errors="coerce")
    except Exception:
        pass
    return df

def signup_upsert(activiteit_id: str, username: str | None, lid_id: str | None,
                  status: str, eating: bool | None, meal_choice: str | None, opmerking: str | None = None):
    if is_readonly():
        raise Exception("Read-only modus")
    assert status in ("yes","no")
    def _lookup(c):
        q = c.table("activity_signups").select("id").eq("activiteit_id", activiteit_id)
        if username: q = q.eq("username", username)
        if lid_id: q = q.eq("lid_id", lid_id)
        return q.limit(1).execute()
    found = run_db(_lookup, what="signups find")
    rows = found.data or []
    payload = {
        "activiteit_id": activiteit_id,
        "status": status,
        "eating": bool(eating) if eating is not None else None,
        "meal_choice": (meal_choice or "").strip() or None,
        "username": username or None,
        "lid_id": lid_id or None,
        "opmerking": (opmerking or "").strip() or None,
        "signup_ts": dt.utcnow().isoformat()
    }
    if rows:
        sid = rows[0]["id"]
        run_db(lambda c: c.table("activity_signups").update(payload).eq("id", sid).execute(), what="signups update")
    else:
        run_db(lambda c: c.table("activity_signups").insert(payload).execute(), what="signups insert")

def appbar(tag: str):
    # BANNER boven menubalk (beheer via Ledenbeheer)
    try:
        _bn_bytes = asset_get_bytes('banner_top')
        _bn_meta = asset_meta_get('banner_top')
    except Exception:
        _bn_bytes, _bn_meta = None, {}
    if _bn_bytes:
        _html = _banner_html(_bn_bytes, _bn_meta)
        if _html:
            st.markdown(_html, unsafe_allow_html=True)
            st.markdown("<div style='height:0.4rem'></div>", unsafe_allow_html=True)
    col1, col2, col3 = st.columns([5, 3, 2])
    with col1:
        st.markdown("**ANWW Duikapp**")
    with col2:
        st.markdown(f"<div class='badge'>{current_username() or '—'} · {current_role()} · Build {APP_BUILD}</div>", unsafe_allow_html=True)
    with col3:
        if st.button("Uitloggen", key=f"logout_{tag}"):
            st.session_state.pop("auth_user", None)
            st.rerun()

def require_role(*allowed):
    if current_role() not in allowed:
        st.error("Onvoldoende rechten."); st.stop()

def page_setup_first_admin_username():
    st.title("Eerste admin aanmaken")
    st.info("Er bestaat nog geen admin. Maak eerst de eerste admin aan.")
    with st.form("first_admin"):
        username = st.text_input("Login (gebruikersnaam)", placeholder="vb. d.verbraeken")
        pw1 = st.text_input("Wachtwoord", type="password")
        pw2 = st.text_input("Herhaal wachtwoord", type="password")
        email = st.text_input("E-mail (optioneel)")
        submitted = st.form_submit_button("Maak admin", type="primary")
    if submitted:
        if not username or not pw1 or len(pw1) < 5 or pw1 != pw2:
            st.error("Controleer login en wachtwoord (min. 5 tekens en gelijk).");
            return
        try:
            auth_create_user(username=username, password=pw1, role="admin", email=email or None)
            payload = {
                "voornaam": "",
                "achternaam": "",
                "email": (email or "").strip().lower() or f"{username}@local",
                "username": username,
                "role": "admin",
                "opt_in_weekly": True,
                "actief": True
            }
            leden_upsert(payload)
            st.success("Admin aangemaakt. Je kan nu inloggen.")
            st.session_state["just_created_admin"] = True
        except Exception as e:
            st.error(f"Mislukt: {e}")

def page_login_username():
    # BANNER op loginpagina (beheer via Ledenbeheer)
    try:
        _lb_bytes = asset_get_bytes('banner_login')
        _lb_meta = asset_meta_get('banner_login')
    except Exception:
        _lb_bytes, _lb_meta = None, {}
    if _lb_bytes:
        _html = _banner_html(_lb_bytes, _lb_meta)
        if _html:
            st.markdown(_html, unsafe_allow_html=True)
            st.markdown("<div style='height:0.6rem'></div>", unsafe_allow_html=True)
    st.title("Inloggen")
    if st.session_state.get("just_created_admin"):
        st.success("Admin aangemaakt. Log nu in."); st.session_state.pop("just_created_admin", None)
    with st.form("login_form"):
        login = st.text_input("Login (gebruikersnaam)")
        pw = st.text_input("Wachtwoord", type="password")
        submitted = st.form_submit_button("Login", type="primary")
    if not submitted: return
    user = auth_get_user_by_username(login)
    if not user or not _verify_password(pw, user.get("password_hash") or ""):
        st.error("Onjuiste login."); return
    st.session_state["auth_user"] = {"id": user.get("id"), "username": user.get("username"), "email": user.get("email"), "role": user.get("role")}
    st.success("Ingelogd."); st.rerun()



def page_profiel():
    appbar("profiel")
    st.header("Mijn profiel")
    uname = current_username()
    if not uname:
        st.info("Niet ingelogd.")
        return
    try:
        row = leden_get_by_username(uname) or {}
    except Exception as e:
        st.error(f"Kon profiel niet laden: {e}")
        return
    # Bepaal beschikbare kolommen
    try:
        _df = leden_list_df()
        _df_cols = set(_df.columns.tolist()) if hasattr(_df, "columns") else set()
    except Exception:
        _df_cols = set(row.keys())

    v_voor   = (row.get("voornaam") or "").strip()
    v_achter = (row.get("achternaam") or "").strip()
    v_email  = (row.get("email") or "").strip().lower()
    v_brevet = normalize_brevet(row.get("duikbrevet"))
    v_tel    = (row.get("telefoon") or "").strip() if ("telefoon" in _df_cols or "telefoon" in row) else ""
    v_adres  = (row.get("adres") or "").strip()    if ("adres" in _df_cols or "adres" in row) else ""
    v_gbd    = row.get("geboortedatum") if ("geboortedatum" in _df_cols or "geboortedatum" in row) else None

    # Parse bestaande telefoon/adres naar losse velden (zoals bij admin ledenbeheer)
    try:
        tel_cc, tel_num = _parse_phone(v_tel) if v_tel else ("+32","")
    except Exception:
        tel_cc, tel_num = ("+32","")
    try:
        adr_parts = _parse_adres_to_parts(v_adres) if v_adres else {"straat":"", "nummer":"", "bus":"", "postcode":"", "stad":"", "land":""}
    except Exception:
        adr_parts = {"straat":"", "nummer":"", "bus":"", "postcode":"", "stad":"", "land":""}

    col1, col2 = st.columns(2)
    with col1:
        i_voor = st.text_input("Voornaam", value=v_voor, key="prof_vn")
        i_achter = st.text_input("Achternaam", value=v_achter, key="prof_an")
        i_email = st.text_input("E-mail", value=v_email, key="prof_em")
        if ("duikbrevet" in _df_cols) or (v_brevet is not None):
            _choices = ["(geen)"] + BREVET_CHOICES
            _cur = v_brevet or "(geen)"
            if _cur not in _choices:
                _cur = "(geen)"
            i_brevet = st.selectbox("Duikbrevet", _choices, index=_choices.index(_cur), key="prof_bv")
        else:
            i_brevet = None
    with col2:
        # Telefoon: landcode + nummer
        if ("telefoon" in _df_cols) or (v_tel is not None):
            st.markdown("**Telefoon**")
            colp1, colp2 = st.columns([0.35, 0.65])
            with colp1:
                i_tel_cc = st.text_input("Landcode", value=(tel_cc or "+32"), key="prof_tel_cc")
            with colp2:
                i_tel_num = st.text_input("Nummer", value=(tel_num or ""), key="prof_tel_num")
        else:
            i_tel_cc = None; i_tel_num = None

        # Adres: straat/nummer/bus + postcode/stad + land
        if ("adres" in _df_cols) or (v_adres is not None):
            st.markdown("**Adres**")
            cola1, cola2, cola3 = st.columns([0.5, 0.25, 0.25])
            with cola1:
                i_adr_straat = st.text_input("Straat", value=adr_parts.get("straat",""), key="prof_adr_straat")
            with cola2:
                i_adr_nummer = st.text_input("Nummer", value=adr_parts.get("nummer",""), key="prof_adr_nummer")
            with cola3:
                i_adr_bus = st.text_input("Bus", value=adr_parts.get("bus",""), key="prof_adr_bus")
            colb1, colb2 = st.columns([0.3, 0.7])
            with colb1:
                i_adr_postcode = st.text_input("Postcode", value=adr_parts.get("postcode",""), key="prof_adr_pc")
            with colb2:
                i_adr_stad = st.text_input("Stad", value=adr_parts.get("stad",""), key="prof_adr_stad")
            i_adr_land = st.text_input("Land", value=(adr_parts.get("land") or "België"), key="prof_adr_land")
        else:
            i_adr_straat = i_adr_nummer = i_adr_bus = i_adr_postcode = i_adr_stad = i_adr_land = None

        # Geboortedatum met bereik 1900-01-01 t/m vandaag
        if ("geboortedatum" in _df_cols) or (v_gbd is not None):
            _default_date = None
            _raw = v_gbd
            if _raw:
                try:
                    import pandas as _pd
                    _dt = _pd.to_datetime(_raw, errors="coerce")
                    if _pd.notna(_dt):
                        _default_date = _dt.date()
                except Exception:
                    _default_date = None
            from datetime import date as _date
            _dmin = _date(1900, 1, 1)
            _dmax = _date.today()
            if _default_date is not None:
                i_gbd = st.date_input("Geboortedatum", value=_default_date, min_value=_dmin, max_value=_dmax, format="DD/MM/YYYY", key="prof_gbd_date")
            else:
                # Als geen waarde bekend is, laat leeg binnen toegestaan bereik
                i_gbd = st.date_input("Geboortedatum", value=None, min_value=_dmin, max_value=_dmax, format="DD/MM/YYYY", key="prof_gbd_date")
            try:
                import pandas as _pd
                _dt_prev = _pd.to_datetime(i_gbd) if i_gbd else None
                if _dt_prev is not None and not _pd.isna(_dt_prev):
                    months = ["januari","februari","maart","april","mei","juni","juli","augustus","september","oktober","november","december"]
                    st.caption(f"Weergave: {int(_dt_prev.day):02d}/{months[int(_dt_prev.month)-1]}/{int(_dt_prev.year)}")
            except Exception:
                pass
        else:
            i_gbd = None

    # Wekelijkse mail: opt-in toggle
    cur_opt = bool(row.get("opt_in_weekly", True))
    i_optin = st.toggle(
        "Ik wil de wekelijkse activiteitenmail ontvangen",
        value=cur_opt,
        key="prof_opt_in_weekly",
    )

    disabled = is_readonly() or not uname
    if st.button("Bewaar mijn gegevens", type="primary", disabled=disabled, key="prof_save_btn"):
        try:
            payload = {
                "username": uname,
                "voornaam": (i_voor or "").strip(),
                "achternaam": (i_achter or "").strip(),
                "email": (i_email or "").strip().lower() or None,
            }
            if i_brevet is not None:
                payload["duikbrevet"] = None if i_brevet == "(geen)" else canon_brevet(i_brevet)

            # Telefoon opslaan als geformatteerde string
            if i_tel_cc is not None or i_tel_num is not None:
                try:
                    payload["telefoon"] = _format_phone_bel_style(i_tel_cc, i_tel_num)
                except Exception:
                    payload["telefoon"] = None

            # Adres samenstellen uit delen
            if i_adr_straat is not None:
                try:
                    payload["adres"] = _format_adres_from_parts(i_adr_straat, i_adr_nummer, i_adr_bus, i_adr_postcode, i_adr_stad, i_adr_land)
                except Exception:
                    payload["adres"] = None

            if i_gbd is not None:
                try:
                    payload["geboortedatum"] = i_gbd.isoformat() if hasattr(i_gbd, "isoformat") else None
                except Exception:
                    payload["geboortedatum"] = None

            # Opt-in voor wekelijkse mail opslaan (indien kolom bestaat)
            payload["opt_in_weekly"] = bool(i_optin)

            leden_upsert(payload)
            st.success("Profiel opgeslagen.")
            st.rerun()
        except Exception as ex:
            st.error(f"Opslaan mislukt: {ex}")
def page_leden():
    """Alle leden tonen (alleen lezen) + export (Excel) voor users & members."""
    st.header("Leden")
    # Ledenlijst ophalen
    try:
        df = leden_list_df()
    except Exception as e:
        st.error(f"Kon ledenlijst niet laden: {e}")
        return
    if df is None or df.empty:
        st.info("Nog geen leden.")
        return

    # Export naar Excel voor iedereen die deze pagina ziet
    export_bytes = None
    try:
        src = df.copy()

        # Helper om veilig kolommen op te halen
        def _col(name: str):
            if name in src.columns:
                return src[name]
            else:
                # zelfde lengte, lege strings
                return pd.Series([""] * len(src))

        export_df = pd.DataFrame({
            "Voornaam": _col("voornaam"),
            "Achternaam": _col("achternaam"),
            "Telefoon": _col("telefoon"),
            "E-mail": _col("email"),
            "Adres": _col("adres"),
            "Geboortedatum": _col("geboortedatum"),
            "Brevet": _col("duikbrevet"),
        })

        # Zet geboortedatum om naar echte datum + dd/mm/jjjj-weergave
        if "Geboortedatum" in export_df.columns:
            gd = pd.to_datetime(export_df["Geboortedatum"], errors="coerce")
            export_df["Geboortedatum"] = gd.dt.date

        sort_cols_exp = [c for c in ["Achternaam", "Voornaam"] if c in export_df.columns]
        if sort_cols_exp:
            export_df = export_df.sort_values(sort_cols_exp, na_position="last")

        buf = io.BytesIO()
        today_str = dt.now().strftime("%d/%m/%Y")
        title = f"Ledenlijst {today_str}"

        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            export_df.to_excel(writer, index=False, sheet_name="Leden", startrow=2)
            ws = writer.sheets["Leden"]

            # Titel in A1 met zelfde stijl als voorbeeld
            ws["A1"] = title
            ws["A1"].font = Font(name="Calibri", bold=True, size=20)
            ws["A1"].alignment = Alignment(horizontal="left")

            # Kolombreedtes zoals voorbeeldbestand
            col_widths = {
                "A": 12.28515625,
                "B": 17.28515625,
                "C": 24.140625,
                "D": 36.28515625,
                "E": 45.0,
                "F": 14.5703125,
                "G": 14.0,
            }
            for col_letter, width in col_widths.items():
                ws.column_dimensions[col_letter].width = width

            # Headerstijl (rij 3): vet + centreren
            header_row_idx = 3
            for cell in ws[header_row_idx]:
                if cell.value is not None:
                    cell.font = Font(name="Calibri", bold=True)
                    cell.alignment = Alignment(horizontal="center")

            # Datum-opmaak voor geboortedatumkolom (dd/mm/jjjj)
            geboortecol = None
            for cell in ws[header_row_idx]:
                if str(cell.value).strip().lower() == "geboortedatum":
                    geboortecol = cell.column_letter
                    break
            if geboortecol:
                max_row = header_row_idx + len(export_df)
                for r in range(header_row_idx + 1, max_row + 1):
                    c = ws[f"{geboortecol}{r}"]
                    if c.value not in (None, ""):
                        c.number_format = "DD/MM/YYYY"

        export_bytes = buf.getvalue()
    except Exception as ex:
        st.error(f"Export mislukt: {ex}")

    if export_bytes:
        fname = f"ledenlijst_{dt.now().strftime('%d%m%Y')}.xlsx"
        st.download_button(
            "⬇️ Exporteer ledenlijst (Excel)",
            data=export_bytes,
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="leden_export_excel"
        )

    # Overzicht op scherm
    role = (current_role() or "").lower()
    if role in ("user", "member"):
        kolommen = [c for c in ["voornaam", "achternaam", "telefoon"] if c in df.columns]
        if not kolommen:
            kolommen = list(df.columns)
    else:
        kolommen = [c for c in ["voornaam", "achternaam", "email", "username", "role"] if c in df.columns]
        if not kolommen:
            kolommen = list(df.columns)

    view = df[kolommen].copy()
    sort_cols = [c for c in ["achternaam", "voornaam"] if c in view.columns]
    if sort_cols:
        view = view.sort_values(sort_cols, na_position="last")
    st.dataframe(view, use_container_width=True, hide_index=True)

    # Detailweergave op naam voor users & members (zonder 'role' en 'username')
    try:
        _role = (current_role() or "").lower()
    except Exception:
        _role = ""
    if _role in ("user", "member"):
        st.caption("Klik op een naam om details te zien.")
        for _row in df.to_dict(orient="records"):
            _fname = (_row.get("voornaam") or "").strip()
            _lname = (_row.get("achternaam") or "").strip()
            _tel = _row.get("telefoon") or ""
            _header = f"{_fname} {_lname}" + (f" — {_tel}" if _tel else "")
            with st.expander(_header):
                for _k, _v in _row.items():
                    if _k in {"role", "username"}:
                        continue
                    st.write(f"**{_k.capitalize()}:** {_v}")


def _parse_adres_to_parts(adres: str):
    import re as _re
    parts = {"straat":"", "nummer":"", "bus":"", "postcode":"", "stad":"", "land":""}
    if not adres: return parts
    s = str(adres).strip()
    chunks = [c.strip() for c in s.split(",") if c.strip()]
    if chunks:
        last = chunks[-1]
        if _re.search(r"[A-Za-z]{2,}", last) and not _re.search(r"\d", last):
            parts["land"] = last; chunks = chunks[:-1]
    pc_idx = None
    for i, c in enumerate(chunks):
        m = _re.match(r"(?P<pc>\d{4,5})\s+(?P<city>.+)$", c)
        if m:
            parts["postcode"] = m.group("pc"); parts["stad"] = m.group("city").strip(); pc_idx = i; break
    if pc_idx is not None: chunks.pop(pc_idx)
    if chunks:
        streetline = chunks[0]
        m2 = _re.match(r"^(?P<street>.+?)\s+(?P<num>\d+)(?:\s*[/\-]?\s*(?P<bus>[A-Za-z0-9]+))?$", streetline)
        if m2:
            parts["straat"] = m2.group("street").strip(); parts["nummer"] = m2.group("num").strip(); parts["bus"] = (m2.group("bus") or "").strip()
        else:
            parts["straat"] = streetline
    return parts

def _format_adres_from_parts(straat, nummer, bus, postcode, stad, land):
    seg1 = " ".join([x for x in [straat, nummer, bus] if x]).strip()
    seg2 = " ".join([x for x in [postcode, stad] if x]).strip()
    segs = [seg for seg in [seg1, seg2, (land or "").strip()] if seg]
    return ", ".join(segs) if segs else None

def _parse_phone(full: str):
    cc, num = "", ""
    if not full: return cc, num
    s = str(full).strip()
    keep_plus = s.startswith("+")
    digits = "".join(ch for ch in s if ch.isdigit())
    if s.startswith("+") or s.startswith("00"):
        for k in (3, 2):
            if len(digits) > k: cc = digits[:k]; num = digits[k:]; break
        if not cc and len(digits) >= 2: cc, num = digits[:2], digits[2:]
    else:
        if len(digits) >= 2: cc, num = digits[:2], digits[2:]
    if keep_plus and cc: cc = "+" + cc
    return cc, num

def _format_phone_bel_style(cc: str, numdigits: str):
    cc = (cc or "").strip()
    if cc and not cc.startswith("+"): cc = "+" + cc
    nd = "".join(ch for ch in str(numdigits or "") if ch.isdigit())
    if not nd: return cc or None
    if len(nd) >= 3:
        a = nd[:3]; rest = nd[3:]; groups = []
        while rest: groups.append(rest[:2]); rest = rest[2:]
        tail = "/" + groups[0] if groups else ""
        if len(groups) > 1: tail += "." + ".".join(groups[1:])
        return f"{cc}(0){a}{tail}"
    else:
        return f"{cc}(0){nd}"

def page_ledenbeheer():
    payload = {}
    require_role("admin")
    if is_readonly(): st.warning("Read-only modus actief — wijzigen uitgeschakeld.")
    st.header("Ledenbeheer (admin)")
    with st.expander('🖼️ Site afbeeldingen (admin)', expanded=False):
        colA, colB = st.columns(2)
        with colA:
            st.markdown('**Banner boven menubalk (JPEG)**')
            _curA = asset_get_bytes('banner_top')
            if _curA: st.image(_curA, use_container_width=True)
            _upA = st.file_uploader('Upload JPEG', type=['jpg','jpeg'], key='upl_banner_top')
            cA1, cA2 = st.columns(2)
            if cA1.button('Bewaar banner', disabled=(is_readonly() or _upA is None)):
                try:
                    asset_set_bytes('banner_top', (_upA.read() if _upA else None), updated_by=(current_username() or current_email()))
                    st.success('Banner opgeslagen.'); st.rerun()
                except Exception as e:
                    st.error(f'Mislukt: {e}')
            if cA2.button('Verwijder banner', disabled=is_readonly()):
                try:
                    asset_set_bytes('banner_top', None, updated_by=(current_username() or current_email()))
                    st.success('Banner verwijderd.'); st.rerun()
                except Exception as e:
                    st.error(f'Mislukt: {e}')
            st.caption('Instellingen (plaats/maat/link)')
            _mA = asset_meta_get('banner_top')
            _wmodeA = st.radio('Breedte', ['Volledige breedte','Aangepast (px)'], index=(1 if (_mA.get('width_mode')=='px') else 0), key='bnA_wmode')
            _wpxA = st.number_input('Breedte (px)', min_value=0, step=10, value=int(_mA.get('width_px') or 0), help='0 = geen vaste breedte', key='bnA_wpx')
            _hpxA = st.number_input('Hoogte (px)', min_value=0, step=10, value=int(_mA.get('height_px') or 0), help='0 = automatische hoogte', key='bnA_hpx')
            _alignA = st.selectbox('Uitlijning', ['links','midden','rechts'], index={'links':0,'midden':1,'rechts':2}.get((_mA.get('align') or 'midden'),1), key='bnA_align')
            _linkA = st.text_input('Link-URL (optioneel)', value=_mA.get('link_url') or '', key='bnA_link')
            if st.button('Bewaar instellingen banner boven', key='bnA_save', disabled=is_readonly()):
                try:
                    metaA = {
                        'width_mode': ('px' if _wmodeA.startswith('Aangepast') else 'full'),
                        'width_px': int(_wpxA or 0),
                        'height_px': int(_hpxA or 0),
                        'align': {'links':'left','midden':'center','rechts':'right'}[_alignA],
                        'link_url': _linkA.strip() or ''
                    }
                    asset_meta_set('banner_top', metaA)
                    st.success('Instellingen opgeslagen.'); st.rerun()
                except Exception as e:
                    st.error(f'Mislukt: {e}')
        with colB:
            st.markdown('**Banner op loginpagina (JPEG)**')
            _curB = asset_get_bytes('banner_login')
            if _curB: st.image(_curB, use_container_width=True)
            _upB = st.file_uploader('Upload JPEG', type=['jpg','jpeg'], key='upl_banner_login')
            cB1, cB2 = st.columns(2)
            if cB1.button('Bewaar login-banner', disabled=(is_readonly() or _upB is None)):
                try:
                    asset_set_bytes('banner_login', (_upB.read() if _upB else None), updated_by=(current_username() or current_email()))
                    st.success('Login-banner opgeslagen.'); st.rerun()
                except Exception as e:
                    st.error(f'Mislukt: {e}')
            if cB2.button('Verwijder login-banner', disabled=is_readonly()):
                try:
                    asset_set_bytes('banner_login', None, updated_by=(current_username() or current_email()))
                    st.success('Login-banner verwijderd.'); st.rerun()
                except Exception as e:
                    st.error(f'Mislukt: {e}')
            st.caption('Instellingen (plaats/maat/link)')
            _mB = asset_meta_get('banner_login')
            _wmodeB = st.radio('Breedte ', ['Volledige breedte','Aangepast (px)'], index=(1 if (_mB.get('width_mode')=='px') else 0), key='bnB_wmode')
            _wpxB = st.number_input('Breedte (px) ', min_value=0, step=10, value=int(_mB.get('width_px') or 0), help='0 = geen vaste breedte', key='bnB_wpx')
            _hpxB = st.number_input('Hoogte (px) ', min_value=0, step=10, value=int(_mB.get('height_px') or 0), help='0 = automatische hoogte', key='bnB_hpx')
            _alignB = st.selectbox('Uitlijning ', ['links','midden','rechts'], index={'links':0,'midden':1,'rechts':2}.get((_mB.get('align') or 'midden'),1), key='bnB_align')
            _linkB = st.text_input('Link-URL (optioneel) ', value=_mB.get('link_url') or '', key='bnB_link')
            if st.button('Bewaar instellingen login-banner', key='bnB_save', disabled=is_readonly()):
                try:
                    metaB = {
                        'width_mode': ('px' if _wmodeB.startswith('Aangepast') else 'full'),
                        'width_px': int(_wpxB or 0),
                        'height_px': int(_hpxB or 0),
                        'align': {'links':'left','midden':'center','rechts':'right'}[_alignB],
                        'link_url': _linkB.strip() or ''
                    }
                    asset_meta_set('banner_login', metaB)
                    st.success('Instellingen opgeslagen.'); st.rerun()
                except Exception as e:
                    st.error(f'Mislukt: {e}')
    df = leden_list_df()
    if not df.empty:
        cols = ["voornaam","achternaam","email","username","role","duikbrevet","opt_in_weekly","actief"]
        show = [c for c in cols if c in df.columns]

        # Export naar Excel vanuit ledenbeheer (admin)
        export_bytes = None
        try:
            src = df.copy()
            def _col(name: str):
                if name in src.columns:
                    return src[name]
                else:
                    return pd.Series([""] * len(src))

            export_df = pd.DataFrame({
                "Voornaam": _col("voornaam"),
                "Achternaam": _col("achternaam"),
                "Telefoon": _col("telefoon"),
                "E-mail": _col("email"),
                "Adres": _col("adres"),
                "Geboortedatum": _col("geboortedatum"),
                "Brevet": _col("duikbrevet"),
            })

            # Zet geboortedatum om naar echte datum + dd/mm/jjjj-weergave
            if "Geboortedatum" in export_df.columns:
                gd = pd.to_datetime(export_df["Geboortedatum"], errors="coerce")
                export_df["Geboortedatum"] = gd.dt.date

            sort_cols_exp = [c for c in ["Achternaam", "Voornaam"] if c in export_df.columns]
            if sort_cols_exp:
                export_df = export_df.sort_values(sort_cols_exp, na_position="last")

            buf = io.BytesIO()
            today_str = dt.now().strftime("%d/%m/%Y")
            title = f"Ledenlijst {today_str}"

            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                export_df.to_excel(writer, index=False, sheet_name="Leden", startrow=2)
                ws = writer.sheets["Leden"]

                ws["A1"] = title
                ws["A1"].font = Font(name="Calibri", bold=True, size=20)
                ws["A1"].alignment = Alignment(horizontal="left")

                col_widths = {
                    "A": 12.28515625,
                    "B": 17.28515625,
                    "C": 24.140625,
                    "D": 36.28515625,
                    "E": 45.0,
                    "F": 14.5703125,
                    "G": 14.0,
                }
                for col_letter, width in col_widths.items():
                    ws.column_dimensions[col_letter].width = width

                header_row_idx = 3
                for cell in ws[header_row_idx]:
                    if cell.value is not None:
                        cell.font = Font(name="Calibri", bold=True)
                        cell.alignment = Alignment(horizontal="center")

                geboortecol = None
                for cell in ws[header_row_idx]:
                    if str(cell.value).strip().lower() == "geboortedatum":
                        geboortecol = cell.column_letter
                        break
                if geboortecol:
                    max_row = header_row_idx + len(export_df)
                    for r in range(header_row_idx + 1, max_row + 1):
                        c = ws[f"{geboortecol}{r}"]
                        if c.value not in (None, ""):
                            c.number_format = "DD/MM/YYYY"

            export_bytes = buf.getvalue()
        except Exception as ex:
            st.error(f"Export mislukt: {ex}")

        if export_bytes:
            fname = f"ledenlijst_{dt.now().strftime('%d%m%Y')}.xlsx"
            st.download_button(
                "⬇️ Exporteer ledenlijst (Excel)",
                data=export_bytes,
                file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="leden_export_excel_admin"
            )

        st.dataframe(df[show].sort_values(["achternaam","voornaam"], na_position="last"),
                     use_container_width=True, hide_index=True)
    else:
        st.info("Nog geen leden.")

    st.divider(); st.subheader("Lid toevoegen")
    with st.form("leden_form", clear_on_submit=True):
        c1, c2, c3 = st.columns(3)
        with c1:
            vn = st.text_input("Voornaam")
            an = st.text_input("Achternaam")
            email = st.text_input("E-mail")
        with c2:
            username = st.text_input("Login (username)*")
            role = st.selectbox("Rol", options=ROLE_CHOICES, index=(ROLE_CHOICES.index("member") if "member" in ROLE_CHOICES else 0))
            brevet = st.selectbox("Duikbrevet", options=["(geen)"] + BREVET_CHOICES, index=0)
        with c3:
            pw1 = st.text_input("Wachtwoord", type="password")
            pw2 = st.text_input("Herhaal wachtwoord", type="password")
            optin = st.toggle("Wekelijkse mail", value=True)
            actief = st.toggle("Actief", value=True)
        submitted = st.form_submit_button("Bewaar lid", type="primary")
        if submitted:
            if not username:
                st.warning("Login (username) is verplicht."); return
            email_eff = (email or "").strip().lower() or f"{username.strip()}@local"
            payload = {
                "email": email_eff,
                "voornaam": (vn or "").strip(),
                "achternaam": (an or "").strip(),
                "username": username.strip(),
                "role": role,
                "duikbrevet": (None if brevet == "(geen)" else canon_brevet(brevet)),
                "opt_in_weekly": bool(optin),
                "actief": bool(actief),
            }
            try:
                leden_upsert(payload)
                # Login aanmaken of bijwerken
                try:
                    u = auth_get_user_by_username(username)
                    if not u:
                        if not pw1 or len(pw1) < 5 or pw1 != pw2:
                            st.warning("Wachtwoord vereist (min. 5 tekens) en beide velden gelijk."); return
                        auth_create_user(username=username, password=pw1, role=role, email=(email or None))
                    else:
                        if not is_readonly():
                            run_db(lambda c: c.table("auth_local").update({"role": role, "email": (email or None)}).eq("username", username).execute(), what="auth_local update")
                except Exception:
                    pass
                st.success("Lid + login bewaard."); st.rerun()
            except Exception as e:
                st.error(f"Bewaren mislukt: {e}")
        st.divider(); st.subheader("Login resetten (nieuw wachtwoord zetten)")
    with st.form("reset_pw"):
        uname = st.text_input("Login (username) van lid"); npw1 = st.text_input("Nieuw wachtwoord", type="password"); npw2 = st.text_input("Herhaal nieuw wachtwoord", type="password")
        ok = st.form_submit_button("Zet nieuw wachtwoord")
    if ok:
        if not uname or not npw1 or len(npw1) < 5 or npw1 != npw2:
            st.warning("Controleer login en wachtwoord (min. 5 tekens, gelijk).")
        else:
            try:
                if not auth_get_user_by_username(uname):
                    st.warning("Er bestaat nog geen login voor deze username. Maak eerst het lid aan (met initieel wachtwoord).");
                else:
                    auth_update_password_by_username(uname, npw1); st.success("Wachtwoord gereset.")
            except Exception as e:
                st.error(f"Reset mislukt: {e}")
    st.divider(); st.subheader("Lid bewerken / verwijderen")
    if 'df' in locals() and df is not None and not df.empty:
        options = []; meta = []; cols = set(df.columns); df_show = df.fillna("")
        for _, r in df_show.iterrows():
            vn = str(r["voornaam"]) if "voornaam" in cols else ""
            an = str(r["achternaam"]) if "achternaam" in cols else ""
            em = str(r["email"]).lower() if "email" in cols else ""
            un = str(r["username"]) if "username" in cols else ""
            rid = r["id"] if "id" in cols else None
            label = f"{an}, {vn}".strip(", ").strip() or (em or un or "(onbekend)")
            info = em or un
            if info: label += f" · {info}"
            options.append(label); meta.append({"id": rid, "email": em, "username": un})
        st.caption("Klik op een lid hieronder om te bewerken/verwijderen:")
        sel = st.radio(" ", options, key="leden_edit_radio")
        if sel:
            i = options.index(sel); m_ = meta[i]
            row = {}
            try:
                if m_.get("id") is not None and "id" in cols:
                    row = df.loc[df["id"] == m_["id"]].iloc[0].to_dict()
                elif m_.get("email") and "email" in cols:
                    row = df.loc[df["email"] == m_["email"]].iloc[0].to_dict()
                elif m_.get("username") and "username" in cols:
                    row = df.loc[df["username"] == m_["username"]].iloc[0].to_dict()
            except Exception:
                row = {}
            with st.form("edit_lid_form", clear_on_submit=False):
                c1, c2 = st.columns(2)
                with c1:
                    vn = st.text_input("Voornaam", value=row.get("voornaam", ""))
                    em = st.text_input("E-mail", value=row.get("email", ""))
                    role = st.selectbox("Rol", ["admin", "user", "member", "viewer"], index=["admin", "user", "member", "viewer"].index((row.get("role") or "member")))
                with c2:
                    an = st.text_input("Achternaam", value=row.get("achternaam", ""))
                    un = st.text_input("Login (username)", value=row.get("username", ""))
                    _choices = ["(geen)"] + (BREVET_CHOICES if "BREVET_CHOICES" in globals() else [])
                    cur_brevet = normalize_brevet(row.get("duikbrevet")) or "(geen)"
                    brevet = st.selectbox("Duikbrevet", _choices, index=_choices.index(cur_brevet) if cur_brevet in _choices else 0)
                optin = st.checkbox("Wekelijkse mail", value=bool(row.get("opt_in_weekly", True)))
                actief = st.checkbox("Actief", value=bool(row.get("actief", True)))
                save = st.form_submit_button("Bewaren", type="primary", disabled=is_readonly())
            if save:
                try:
                    payload = {
                        "email": (em or "").strip().lower(),
                        "voornaam": (vn or "").strip(),
                        "achternaam": (an or "").strip(),
                        "username": (un or "").strip(),
                        "role": (role or "member").strip(),
                        "duikbrevet": (None if brevet == "(geen)" else canon_brevet(brevet)),
                        "opt_in_weekly": bool(optin),
                        "actief": bool(actief),
                    }
                    leden_upsert(payload); st.success("Lid bewaard."); st.rerun()
                except Exception as e:
                    st.error(f"Bewaren mislukt: {e}")
            del_disabled = is_readonly() or not any([m_.get("id"), m_.get("email"), m_.get("username")])
            if st.button("Verwijder lid", type="secondary", disabled=del_disabled, key="leden_delete_btn"):
                try:
                    if m_.get("id") and "id" in cols:
                        run_db(lambda c: c.table("leden").delete().eq("id", m_["id"]).execute(), what="leden delete by id")
                    elif m_.get("email") and "email" in cols:
                        run_db(lambda c: c.table("leden").delete().eq("email", m_["email"]).execute(), what="leden delete by email")
                    elif m_.get("username") and "username" in cols:
                        run_db(lambda c: c.table("leden").delete().eq("username", m_["username"]).execute(), what="leden delete by username")
                    st.success("Lid verwijderd."); st.rerun()
                except Exception as e:
                    st.error(f"Verwijderen mislukt: {e}")
    else:
        st.info("Nog geen leden om te bewerken.")

def page_activiteiten():
    appbar("activiteiten"); st.header("Kalender & Inschrijvingen")
    if current_role() in {"admin","user","viewer"}:
        with st.expander("➕ Nieuwe activiteit"):
            # Reset de create-activiteit velden bij een nieuwe run (na succesvol bewaren)
            if st.session_state.get("_reset_act", False):
                for _k in ["act_title","act_omschr","act_notes","act_datum","act_time_tp","act_time_tw","act_loc_select","act_new_loc","act_meal1","act_meal2","act_meal3"]:
                    try:
                        st.session_state.pop(_k, None)
                    except Exception:
                        pass
                st.session_state["_reset_act"] = False
            c1, c2 = st.columns([2, 1])
            with c1:
                titel = st.text_input("Titel*", key="act_title"); omschr = st.text_area("Omschrijving", key="act_omschr"); opmerkingen = st.text_area("Opmerkingen (optioneel)", key="act_notes")
            with c2:
                datum = st.date_input("Datum*", value=datetime.date.today(), format="DD/MM/YYYY", key="act_datum")
                tijd_tp = st.time_input("Ter plaatse (optioneel)", value=None, key="act_time_tp")
                tijd_tw = st.time_input("Te water (optioneel)", value=None, key="act_time_tw")
                tijd = tijd_tp
                pl = plaatsen_list()
                locatie = st.selectbox("Locatie", ["— kies —"] + pl, index=0, key="act_loc_select")
                new_loc = st.text_input("Nieuwe locatie (indien niet in lijst)", key="act_new_loc")
                if st.button("➕ Locatie toevoegen", key="act_add_loc_btn", disabled=is_readonly()):
                    if new_loc and new_loc not in pl:
                        try: plaats_add(new_loc); st.success("Locatie toegevoegd. Kies nu uit de lijst."); st.rerun()
                        except Exception as e: st.error(f"Mislukt: {e}")
                    else: st.warning("Leeg of al bestaand.")
            # Organisator (keuze uit leden)
            try:
                _ld = leden_list_df()
                _ld = _ld.dropna(subset=['username']).copy() if isinstance(_ld, pd.DataFrame) else pd.DataFrame([])
            except Exception:
                _ld = pd.DataFrame([])
            _org_labels, _org_map = ['— kies —'], {'— kies —': None}
            if not _ld.empty:
                for _, _r in _ld.sort_values(['voornaam','achternaam']).iterrows():
                    _u = str(_r.get('username') or '').strip()
                    if not _u: continue
                    _label = f"{_r.get('voornaam','').strip()} {_r.get('achternaam','').strip()} ({_u})".strip()
                    _org_labels.append(_label); _org_map[_label] = _u
            organisator_sel = st.selectbox('Organisator', _org_labels, index=0, key='act_org')
            organisator_user = _org_map.get(organisator_sel)

            st.caption("Maaltijdopties (max. 3, optioneel)")
            m1, m2, m3 = st.columns(3)
            with m1: mo1 = st.text_input("Optie 1", key="act_meal1")
            with m2: mo2 = st.text_input("Optie 2", key="act_meal2")
            with m3: mo3 = st.text_input("Optie 3", key="act_meal3")
            if st.button("Activiteit toevoegen", type="primary", key="act_add_btn", disabled=is_readonly()):
                if not titel or not datum:
                    st.warning("Titel en datum zijn verplicht.")
                else:
                    meal_opts = [x.strip() for x in [mo1, mo2, mo3] if x and x.strip()]
                    if tijd_tw:
                        _tw = tijd_tw.strftime('%H:%M')
                        omschr = (omschr or '')
                        if 'Te water:' not in omschr:
                            omschr = (omschr + ('\n' if omschr else '') + f'Te water: {_tw}')
                    try:
                        activiteit_add(titel=titel, omschr=omschr, datum=datum, tijd=tijd,
                                       locatie=None if not locatie or locatie == "— kies —" else locatie.strip(),
                                       meal_opts=meal_opts or None, created_by=current_username() or current_email(),
                                       opmerkingen=opmerkingen or None, organisator=organisator_user)
                        # Veldreset markeren voor volgende run
                        st.session_state["_reset_act"] = True
                        st.success("Activiteit aangemaakt."); st.rerun()
                    except Exception as e:
                        st.error(f"Mislukt: {e}")
    df = activiteiten_list_df(upcoming=True)
    if df.empty: st.info("Geen (toekomstige) activiteiten."); return
    my_username = current_username(); my_lid = leden_get_by_username(my_username); my_lid_id = (my_lid or {}).get("id")
    for _, row in df.sort_values(["datum", "tijd"], na_position="last").iterrows():
        s = signups_get(row["id"]); myrow = None
        try:
            leden_df = leden_list_df()
            _ldu = leden_df[['username','voornaam','achternaam']].dropna(how='all') if isinstance(leden_df, pd.DataFrame) else None
            if isinstance(s, pd.DataFrame) and not s.empty and _ldu is not None and not _ldu.empty:
                s = s.merge(_ldu, on='username', how='left')
                if 'voornaam' not in s.columns or s['voornaam'].isna().all():
                    _ldi = leden_df[['id','voornaam','achternaam']]
                    s = s.merge(_ldi, left_on='lid_id', right_on='id', how='left', suffixes=('', '_lid'))
        except Exception:
            pass
        if my_username:
            tmp = s.loc[s["username"] == my_username]
            if not tmp.empty: myrow = tmp
        if (myrow is None or myrow.empty) and my_lid_id:
            tmp = s.loc[s["lid_id"] == my_lid_id]
            if not tmp.empty: myrow = tmp
        my_status = (myrow.iloc[0]["status"] if (myrow is not None and not myrow.empty) else None)
        badge = "🟢 ingeschreven" if my_status == "yes" else ("🔴 niet ingeschreven" if my_status == "no" else "⚪ nog niet gekozen")
        titel = f"{row['titel']} — {pd.to_datetime(row['datum']).strftime('%d/%m/%Y')}"
        # Organisator achter titel (indien aanwezig)
        try:
            _org_u = row.get('organisator')
            _org_lbl = None
            if _org_u:
                _ldn = leden_list_df()
                if isinstance(_ldn, pd.DataFrame) and not _ldn.empty:
                    _one = _ldn.loc[_ldn['username'] == _org_u]
                    if not _one.empty:
                        _org_lbl = f"{_one.iloc[0].get('voornaam','').strip()} {_one.iloc[0].get('achternaam','').strip()}"
            if _org_lbl:
                titel += f" — {_org_lbl}"
        except Exception:
            pass
        tstr = None
        if row.get('tijd'):
            _t_raw = str(row.get('tijd'))
            try:
                _t = pd.to_datetime(_t_raw, errors='coerce')
                if pd.notna(_t):
                    tstr = _t.strftime('%H:%M')
            except Exception:
                tstr = None
            if tstr is None:
                import re as _re
                m = _re.match(r'^(\d{1,2}:\d{2})(?::\d{2})?$', _t_raw)
                if m: tstr = m.group(1)
        if tstr: titel += f' · {tstr}'
        # Tellers (rechts) — inclusief extra mee-eters
        yes_count = int((s['status'] == 'yes').sum()) if isinstance(s, pd.DataFrame) and not s.empty else 0
        no_count  = int((s['status'] == 'no').sum())  if isinstance(s, pd.DataFrame) and not s.empty else 0
        extra_sum = 0
        if isinstance(s, pd.DataFrame) and not s.empty and 'opmerking' in s.columns:
            import re as _re
            def _extra_from(_x):
                if not isinstance(_x, str): return 0
                m = _re.search(r'extra\s*mee-eters\s*:\s*(\d+)', _x, _re.I)
                return int(m.group(1)) if m else 0
            try:
                extra_sum = int(s['opmerking'].apply(_extra_from).sum())
            except Exception:
                extra_sum = 0
        tot_yes = yes_count + extra_sum
        col_main, col_yes, col_no = st.columns([14, 1, 1])
        with col_yes:
            st.markdown(f"<div style='background:#3CA133;color:#fff;border:2px solid var(--border);border-radius:8px;padding:0.1rem 0.2rem;text-align:center;min-width:2.2rem;font-weight:700;'>{tot_yes}</div>", unsafe_allow_html=True)
        with col_no:
            st.markdown(f"<div style='background:#ef4444;color:#fff;border:2px solid var(--border);border-radius:8px;padding:0.1rem 0.2rem;text-align:center;min-width:2.2rem;font-weight:700;'>{no_count}</div>", unsafe_allow_html=True)
        with col_main:
            with st.expander(f"{titel}   ·   {badge}", expanded=False):
                if row.get("locatie"): st.caption(f"📍 {row['locatie']}")
                if row.get("omschrijving"): st.write(row["omschrijving"])
                if row.get("opmerkingen"): st.write(row["opmerkingen"])
                coming = s.loc[s["status"] == "yes"].sort_values("signup_ts"); notcoming = s.loc[s["status"] == "no"].sort_values("signup_ts")
                colA, colB = st.columns(2)
                with colA:
                    st.markdown("**Komen (op volgorde van inschrijving):**")
                    if coming.empty: st.caption("Nog niemand.")
                    else:
                        for _, ss in coming.iterrows():
                            meal = f" · eet: {ss['meal_choice']}" if ss.get("eating") else ""
                            _vn, _an = ss.get('voornaam'), ss.get('achternaam')
                            _nm = (f"{_vn} {_an}".strip() if (isinstance(_vn, str) or isinstance(_an, str)) else (ss.get('username') or 'lid'))
                            _ts = pd.to_datetime(ss.get('signup_ts'), errors='coerce')
                            _tstr = _ts.strftime('%d/%m/%Y %H:%M') if pd.notna(_ts) else ''
                            st.write(f"- {_nm}{(' · ' + _tstr) if _tstr else ''}{meal}")
                with colB:
                    st.markdown("**Niet komen:**")
                    if notcoming.empty: st.caption("Nog niemand.")
                    else:
                        for _, ss in notcoming.iterrows():
                            _vn, _an = ss.get('voornaam'), ss.get('achternaam')
                            _nm = (f"{_vn} {_an}".strip() if (isinstance(_vn, str) or isinstance(_an, str)) else (ss.get('username') or 'lid'))
                            _ts = pd.to_datetime(ss.get('signup_ts'), errors='coerce')
                            _tstr = _ts.strftime('%d/%m/%Y %H:%M') if pd.notna(_ts) else ''
                            st.write(f"- {_nm}{(' · ' + _tstr) if _tstr else ''}")
                    # --- Inschrijvingen beheren (alleen admin/user) ---
                    if current_role() in {"admin","user"}:
                        with st.expander("👥 Inschrijvingen beheren", expanded=False):
                            leden_df = leden_list_df()
                            try:
                                s_all = signups_get(row["id"])
                            except Exception:
                                s_all = pd.DataFrame([])
                            # Toevoegen
                            existing_users = set(s_all["username"].dropna().tolist()) if not s_all.empty and "username" in s_all.columns else set()
                            user_options = [u for u in sorted(leden_df["username"].dropna().unique().tolist()) if u not in existing_users]
                            add_user = st.selectbox("Lid toevoegen", ["— kies —"] + user_options, key=f"addu_{row['id']}")
                            add_status = st.radio("Status", ["Komt", "Komt niet"], horizontal=True, key=f"addst_{row['id']}")
                            add_eat = st.checkbox("Eet mee", value=False, key=f"addeat_{row['id']}")
                            meal_opts = row.get("meal_options") or []
                            add_meal = None
                            if add_eat and meal_opts:
                                add_meal_sel = st.selectbox("Maaltijd", ["— kies —"] + meal_opts, index=0, key=f"addmeal_{row['id']}")
                                add_meal = None if add_meal_sel == "— kies —" else add_meal_sel
                            add_extra = st.number_input("Extra mee-eters", min_value=0, max_value=50, step=1, value=0, key=f"addextra_{row['id']}") if add_eat and add_status == "Komt" else 0
                            if st.button("Toevoegen", key=f"addbtn_{row['id']}", disabled=(add_user == "— kies —" or is_readonly())):
                                try:
                                    opm = (f"[extra mee-eters: {add_extra}]" if add_extra else None)
                                    signup_upsert(activiteit_id=row["id"], username=add_user, lid_id=None,
                                                  status=("yes" if add_status == "Komt" else "no"), eating=(add_eat if add_status == "Komt" else None),
                                                  meal_choice=(add_meal if add_eat and add_status == "Komt" else None), opmerking=opm)
                                    st.success("Lid toegevoegd / bijgewerkt."); st.rerun()
                                except Exception as ex:
                                    st.error(f"Toevoegen mislukt: {ex}")
                            # Bezoeker toevoegen
                            st.markdown("**Bezoeker toevoegen**")
                            vcol1, vcol2 = st.columns(2)
                            with vcol1:
                                vis_vn = st.text_input("Voornaam (bezoeker)", key=f"vis_vn_{row['id']}")
                                vis_an = st.text_input("Naam (bezoeker)", key=f"vis_an_{row['id']}")
                            with vcol2:
                                vis_tel = st.text_input("Telefoon (bezoeker)", key=f"vis_tel_{row['id']}")
                                vis_brevet = st.text_input("Brevet (bezoeker)", key=f"vis_brv_{row['id']}")
                            can_add_vis = bool((vis_vn or "").strip() and (vis_an or "").strip())
                            if st.button("Bezoeker toevoegen", key=f"addvis_{row['id']}", disabled=(not can_add_vis or is_readonly())):
                                try:
                                    disp_name = f"{(vis_vn or '').strip()} {(vis_an or '').strip()} (bezoeker)"
                                    extra_parts = []
                                    if vis_tel and str(vis_tel).strip():
                                        extra_parts.append(f"tel: {str(vis_tel).strip()}")
                                    if vis_brevet and str(vis_brevet).strip():
                                        extra_parts.append(f"brevet: {str(vis_brevet).strip()}")
                                    opm_vis = None
                                    if extra_parts:
                                        opm_vis = "[bezoeker] " + ", ".join(extra_parts)
                                    signup_upsert(
                                        activiteit_id=row["id"],
                                        username=disp_name,
                                        lid_id=None,
                                        status="yes",
                                        eating=None,
                                        meal_choice=None,
                                        opmerking=opm_vis,
                                    )
                                    st.success("Bezoeker toegevoegd."); st.rerun()
                                except Exception as ex:
                                    st.error(f"Toevoegen bezoeker mislukt: {ex}")
                            st.markdown("---")
                            # Lijst van ingeschrevenen
                            if s_all is None or s_all.empty:
                                st.caption("Nog geen inschrijvingen.")
                            else:
                                for _, srow in s_all.sort_values(by=["signup_ts"], ascending=[True]).iterrows():
                                    uname = srow.get("username") or "lid"
                                    c1, c2, c3, c4, c5 = st.columns([2,1,1,2,1])
                                    with c1:
                                        st.write(uname)
                                        try:
                                            _ts = pd.to_datetime(srow.get('signup_ts'), errors='coerce')
                                            if pd.notna(_ts):
                                                st.caption(_ts.strftime('%d/%m/%Y %H:%M'))
                                        except Exception:
                                            pass
                                    with c2:
                                        st.write("Status"); st.session_state.setdefault(f"edst_{row['id']}_{uname}", "Komt" if srow.get("status") == "yes" else "Komt niet")
                                        edst = st.selectbox("", ["Komt","Komt niet"], key=f"edst_{row['id']}_{uname}")
                                    with c3:
                                        st.write("Eet mee"); st.session_state.setdefault(f"edeat_{row['id']}_{uname}", bool(srow.get("eating")) if pd.notna(srow.get("eating")) else False)
                                        edeat = st.checkbox("", value=st.session_state[f"edeat_{row['id']}_{uname}"], key=f"edeat_{row['id']}_{uname}")
                                    with c4:
                                        meal_opts = row.get("meal_options") or []
                                        prev_meal = srow.get("meal_choice") if isinstance(srow.get("meal_choice"), str) and srow.get("meal_choice").strip() else None
                                        def_ix = 0
                                        if prev_meal and prev_meal in meal_opts: def_ix = meal_opts.index(prev_meal) + 1
                                        emeal_sel = st.selectbox("Maaltijd", ["— kies —"] + meal_opts, index=def_ix, key=f"emeal_{row['id']}_{uname}") if edeat and meal_opts and edst == "Komt" else None
                                    with c5:
                                        if st.button("Opslaan", key=f"edsv_{row['id']}_{uname}", disabled=is_readonly()):
                                            try:
                                                signup_upsert(activiteit_id=row["id"], username=(None if uname == 'lid' else uname), lid_id=None,
                                                              status=("yes" if edst == "Komt" else "no"), eating=(edeat if edst == "Komt" else None),
                                                              meal_choice=(emeal_sel if edeat and edst == "Komt" else None), opmerking=None)
                                                st.success("Opgeslagen."); st.rerun()
                                            except Exception as ex:
                                                st.error(f"Opslaan mislukt: {ex}")
                                        delcol = st.button("Verwijder", key=f"eddel_{row['id']}_{uname}", disabled=is_readonly())
                                    if delcol:
                                        try:
                                            run_db(signups_delete_users(row["id"], usernames=[uname]), what="signups delete (admin)")
                                            st.success("Verwijderd."); st.rerun()
                                        except Exception as ex:
                                            st.error(f"Verwijderen mislukt: {ex}")
                    # --- einde beheer ---

                if current_role() in {"admin", "user", "member"} and not is_readonly():
                    st.divider(); st.markdown("**Mijn inschrijving**")
                    prev_eating, prev_meal, prev_remark = False, None, None
                    if myrow is not None and not myrow.empty:
                        if pd.notna(myrow.iloc[0].get('eating')): prev_eating = bool(myrow.iloc[0].get('eating'))
                        pm = myrow.iloc[0].get('meal_choice'); prev_meal = pm if isinstance(pm, str) and pm.strip() else None
                        pr = myrow.iloc[0].get('opmerking'); prev_remark = pr if isinstance(pr, str) else None
                    prev_extra = 0
                    if isinstance(prev_remark, str):
                        try:
                            import re as _re
                            m = _re.search(r'extra\s*mee-eters\s*:\s*(\d+)', prev_remark, _re.I)
                            if m: prev_extra = int(m.group(1))
                        except Exception:
                            prev_extra = 0
                    
                    # Vinkvakjes: Ik kom / Ik kom niet
                    yes_ck = st.checkbox('Ik kom', value=(my_status=='yes'), key=f'ck_yes_{row["id"]}')
                    no_ck  = st.checkbox('Ik kom niet', value=(my_status=='no'), key=f'ck_no_{row["id"]}')
                    if yes_ck and no_ck:
                        # beide aangevinkt -> geen geldige keuze
                        status = '— kies —'
                    elif yes_ck:
                        status = 'Ik kom'
                    elif no_ck:
                        status = 'Ik kom niet'
                    else:
                        status = '— kies —'
                    
                    # Eet mee (alleen relevant als je komt)
                    eating_val = st.checkbox('Ik eet mee', value=prev_eating, key=f'eat_{row["id"]}')
                    eating = eating_val if status == 'Ik kom' else None
                    
                    # Extra mee-eters (persistent via session_state + voorvulling uit opmerking)
                    extra_key = f'eat_extra_{row["id"]}'
                    default_extra = int(st.session_state.get(extra_key, prev_extra))
                    if status == 'Ik kom' and eating:
                        extra_people = st.number_input('Extra mee-eters', min_value=0, max_value=50, step=1, value=default_extra, key=extra_key)
                    else:
                        extra_people = int(st.session_state.get(extra_key, default_extra))
                    
                    # Maaltijdkeuze (indien van toepassing)
                    meal_choice = None; meal_opts = row.get('meal_options') or []
                    if status == 'Ik kom' and eating and meal_opts:
                        default_ix = 0
                        if prev_meal and prev_meal in meal_opts: default_ix = meal_opts.index(prev_meal) + 1
                        mc = st.selectbox('Kies je maaltijd', ['— kies —'] + meal_opts, index=default_ix, key=f'meal_{row["id"]}')
                        meal_choice = None if mc == '— kies —' else mc
                    
                    # Opmerking
                    remark_val = st.text_area('Opmerking (optioneel)', value=(prev_remark or ''), key=f'remark_{row["id"]}')
                    
                    if st.button('Bewaar mijn keuze', key=f'save_{row["id"]}', type='primary', disabled=(status == '— kies —')):
                        try:
                            # Strip oude [extra mee-eters: N] uit opmerking en voeg de actuele weer toe
                            import re as __re
                            base_remark = __re.sub(r'\s*\[extra\s*mee-eters\s*:\s*\d+\]\s*', '', (remark_val or ''), flags=__re.I).strip()
                            opm_final = (base_remark + (f' [extra mee-eters: {extra_people}]' if extra_people else '')).strip() if status == 'Ik kom' else None
                            signup_upsert(activiteit_id=row['id'], username=(my_username or None), lid_id=(my_lid_id or None),
                                          status=('yes' if status == 'Ik kom' else 'no'), eating=eating, meal_choice=meal_choice, opmerking=opm_final)
                            st.success('Inschrijving bijgewerkt.'); st.rerun()
                        except Exception as e:
                            st.error(f'Opslaan mislukt: {e}')
                            # --- Beheer activiteit (admin & user) ---
                            if current_role() in {"admin","user"}:
                                with st.expander("⚙️ Beheer activiteit", expanded=(st.session_state.get('open_edit_id') == row['id'])):
                                    st.markdown("**Bewerk activiteit**")
                                    ec1, ec2 = st.columns([2,1])
                                    with ec1:
                                        etitel = st.text_input("Titel*", value=(row.get("titel") or ""), key=f"e_title_{row['id']}")
                                        eomschr = st.text_area("Omschrijving", value=(row.get("omschrijving") or ""), key=f"e_om_{row['id']}")
                                        eopm = st.text_area("Opmerkingen (optioneel)", value=(row.get("opmerkingen") or ""), key=f"e_opm_{row['id']}")
                                    with ec2:
                                        try:
                                            _d = pd.to_datetime(row.get("datum")).date() if row.get("datum") else datetime.date.today()
                                        except Exception:
                                            _d = datetime.date.today()
                                        try:
                                            _t = pd.to_datetime(row.get("tijd")).time() if row.get("tijd") else None
                                        except Exception:
                                            _t = None
                                        edatum = st.date_input("Datum*", value=_d, format="DD/MM/YYYY", key=f"e_date_{row['id']}")
                                        etijd = st.time_input("Tijd (optioneel)", value=_t, key=f"e_time_{row['id']}")
                                        _pl = plaatsen_list()
                                        _loc = row.get("locatie") or ""
                                        _ix = 1 + _pl.index(_loc) if _loc in _pl else 0
                                        eloc = st.selectbox("Locatie", ["— kies —"] + _pl, index=_ix, key=f"e_loc_{row['id']}")
                                    # Organisator (bewerken)
                                    try:
                                        _ld2 = leden_list_df()
                                        _ld2 = _ld2.dropna(subset=['username']).copy() if isinstance(_ld2, pd.DataFrame) else pd.DataFrame([])
                                    except Exception:
                                        _ld2 = pd.DataFrame([])
                                    _org_labels2, _org_map2 = ['— kies —'], {'— kies —': None}
                                    _cur_org = row.get('organisator') or None
                                    _def_ix = 0
                                    if not _ld2.empty:
                                        _ld2 = _ld2.sort_values(['voornaam','achternaam'])
                                        for _ix2, _r2 in _ld2.iterrows():
                                            _u2 = str(_r2.get('username') or '').strip()
                                            if not _u2: continue
                                            _lbl2 = f"{_r2.get('voornaam','').strip()} {_r2.get('achternaam','').strip()} ({_u2})".strip()
                                            _org_labels2.append(_lbl2); _org_map2[_lbl2] = _u2
                                            if _u2 == _cur_org: _def_ix = len(_org_labels2) - 1
                                    organisator_sel2 = st.selectbox('Organisator', _org_labels2, index=_def_ix, key=f"e_org_{row['id']}")
                                    organisator_user2 = _org_map2.get(organisator_sel2)

                                    em1, em2, em3 = st.columns(3)
                                    _opts = row.get("meal_options") or []
                                    with em1: emo1 = st.text_input("Optie 1", value=(_opts[0] if len(_opts)>0 else ""), key=f"e_m1_{row['id']}")
                                    with em2: emo2 = st.text_input("Optie 2", value=(_opts[1] if len(_opts)>1 else ""), key=f"e_m2_{row['id']}")
                                    with em3: emo3 = st.text_input("Optie 3", value=(_opts[2] if len(_opts)>2 else ""), key=f"e_m3_{row['id']}")
                                    if st.button("Bewaar wijzigingen", type="primary", key=f"e_save_{row['id']}", disabled=is_readonly()):
                                        _meal = [x.strip() for x in [emo1, emo2, emo3] if x and x.strip()]
                                        payload = {
                                            "titel": etitel.strip(),
                                            "omschrijving": (eomschr or "").strip(),
                                            "datum": edatum.isoformat(),
                                            "tijd": (etijd.isoformat() if etijd else None),
                                            "locatie": (None if eloc == "— kies —" else eloc.strip()),
                                            "meal_options": (_meal or None),
                                            "opmerkingen": (eopm or "").strip() or None,
                                            "organisator": (organisator_user2 or None)
                                        }
                                        try:
                                            activiteit_update(row["id"], payload)
                                            st.success("Activiteit bijgewerkt."); st.rerun()
                                        except Exception as ex:
                                            st.error(f"Opslaan mislukt: {ex}")
                                    st.markdown("---")
                                    st.markdown("**Deelnemers beheren**")
                                    try:
                                        s_all = signups_get(row["id"])
                                    except Exception:
                                        s_all = pd.DataFrame([])
                                    leden_df = leden_list_df()
                                    user_options = sorted([r.get("username") for _, r in leden_df.iterrows() if r.get("username")])
                                    add_users = st.multiselect("Leden toevoegen (komen)", user_options, key=f"e_add_{row['id']}")
                                    if st.button("Voeg leden toe", key=f"e_add_btn_{row['id']}", disabled=(is_readonly() or not add_users)):
                                        try:
                                            for _u in add_users:
                                                signup_upsert(activiteit_id=row["id"], username=_u, lid_id=None, status="yes", eating=None, meal_choice=None)
                                            st.success(f"Toegevoegd: {len(add_users)}"); st.rerun()
                                        except Exception as ex:
                                            st.error(f"Toevoegen mislukt: {ex}")
                                    cur_users = []
                                    if not s_all.empty:
                                        try:
                                            cur_users = sorted([u for u in s_all[s_all["status"]=="yes"]["username"].dropna().unique().tolist()])
                                        except Exception:
                                            cur_users = []
                                    rem_users = st.multiselect("Leden verwijderen (uit activiteit)", cur_users, key=f"e_rem_{row['id']}")
                                    if st.button("Verwijder geselecteerden", key=f"e_rem_btn_{row['id']}", disabled=(is_readonly() or not rem_users)):
                                        try:
                                            run_db(signups_delete_users(row["id"], usernames=rem_users), what="signups delete (admin)")
                                            st.success(f"Verwijderd: {len(rem_users)}"); st.rerun()
                                        except Exception as ex:
                                            st.error(f"Verwijderen mislukt: {ex}")
                            # --- Einde adminblok ---

            

                # --- Export naar Excel voor deze activiteit ---


            

                try:


            

                    # Meta (bovenbalk)


            

                    _d = pd.to_datetime(row.get("datum"), errors="coerce")


            

                    _datum_str = _d.strftime("%d/%m/%Y") if pd.notna(_d) else ""


            

                    _titel_x = row.get("titel") or ""


            

                    _loc_x = row.get("locatie") or ""


            

                    # Te water (uit omschrijving)


            

                    import re as __re


            

                    _om = row.get("omschrijving") or ""


            

                    _tw = ""


            

                    try:


            

                        _m = __re.search(r"Te\s*water\s*:\s*(\d{1,2}:\d{2})", _om, __re.I)


            

                        if _m: _tw = _m.group(1)


            

                    except Exception:


            

                        _tw = ""


            

                    # Organisator als 'Voornaam Achternaam (username)' indien beschikbaar


            

                    _org_u = row.get("organisator") or ""


            

                    _org_disp = _org_u


            

                    try:


            

                        _leden_df = leden_list_df()


            

                        if _org_u and isinstance(_leden_df, pd.DataFrame) and not _leden_df.empty and "username" in _leden_df.columns:


            

                            _r = _leden_df.loc[_leden_df["username"] == _org_u]


            

                            if not _r.empty:


            

                                _vn0 = (_r.iloc[0].get("voornaam") or "").strip()


            

                                _an0 = (_r.iloc[0].get("achternaam") or "").strip()


            

                                _org_disp = f"{_vn0} {_an0} ({_org_u})".strip()


            

                    except Exception:


            

                        pass


            

                    _meta_df = pd.DataFrame([{"Datum": _datum_str, "Titel": _titel_x, "Locatie": _loc_x, "Te water": _tw, "Organisator": _org_disp}])


            

                


            

                    # Deelnemers die ingeschreven hebben (status == yes)


            

                    _s_base = s.copy() if isinstance(s, pd.DataFrame) else pd.DataFrame([])


            

                    if not _s_base.empty and "status" in _s_base.columns:


            

                        _s_base = _s_base[_s_base["status"] == "yes"].copy()


            

                    else:


            

                        _s_base = pd.DataFrame([])


            

                


            

                    # Merge met leden voor namen en brevet


            

                    try:


            

                        _leden = leden_list_df()


            

                    except Exception:


            

                        _leden = pd.DataFrame([])


            

                


            

                    _s1 = _s_base.copy()


            

                    if not _s1.empty and isinstance(_leden, pd.DataFrame) and not _leden.empty:


            

                        # via username (indien beschikbaar)


            

                        if "username" in _s1.columns and "username" in _leden.columns:


            

                            _keep1 = [c for c in ["username","voornaam","achternaam","duikbrevet"] if c in _leden.columns]


            

                            _s1 = _s1.merge(_leden[_keep1], on="username", how="left")


            

                        # via lid_id (fallback)


            

                        if "lid_id" in _s1.columns and "id" in _leden.columns:


            

                            _keep2 = [c for c in ["id","voornaam","achternaam","duikbrevet"] if c in _leden.columns]


            

                            _s1 = _s1.merge(_leden[_keep2].rename(columns={"id":"lid_id","voornaam":"voornaam_by_id","achternaam":"achternaam_by_id","duikbrevet":"duikbrevet_by_id"}), on="lid_id", how="left")


            

                


            

                    # Kolommen opbouwen met veilige fallbacks


            

                    _N = len(_s1) if isinstance(_s1, pd.DataFrame) else 0


            

                    _empty = (pd.Series([None]*_N) if _N else pd.Series([], dtype=object))


            

                    _vn = _s1.get("voornaam", _empty)


            

                    _vn = _vn.combine_first(_s1.get("voornaam_by_id", _empty)) if _N else _vn


            

                    _an = _s1.get("achternaam", _empty)


            

                    _an = _an.combine_first(_s1.get("achternaam_by_id", _empty)) if _N else _an


            

                    _brev = _s1.get("duikbrevet", _empty)


            

                    _brev = _brev.combine_first(_s1.get("duikbrevet_by_id", _empty)) if _N else _brev


            

                


            

                    # Eet mee / maaltijd / extra mee-eters


            

                    _eat = _s1.get("eating", pd.Series([False]*_N) if _N else pd.Series([], dtype=bool)).apply(lambda v: bool(v) if pd.notna(v) else False) if _N else pd.Series([], dtype=bool)


            

                    _meal = _s1.get("meal_choice", pd.Series([""]*_N) if _N else pd.Series([], dtype=object))


            

                    def __extra_people(x):


            

                        try:


            

                            import re as ___re


            

                            m = ___re.search(r"extra\s*mee-eters\s*:\s*(\d+)", str(x or ""), ___re.I)


            

                            return int(m.group(1)) if m else 0


            

                        except Exception:


            

                            return 0


            

                    _extra = _s1.get("opmerking", pd.Series([""]*_N) if _N else pd.Series([], dtype=object)).apply(__extra_people) if _N else pd.Series([], dtype=int)


            

                


            

                    # Normaliseer brevetten + sorteer op brevet, achternaam, voornaam


            

                    try:


            

                        _brev = _brev.apply(canon_brevet) if _N else _brev


            

                    except Exception:


            

                        pass


            

                    _part_df = pd.DataFrame({


            

                        "Voornaam": _vn, "Achternaam": _an, "Brevet": _brev,


            

                        "Eet mee": _eat, "Maaltijd": _meal, "Extra mee-eters": _extra


            

                    })


            

                    if not _part_df.empty:


            

                        try:


            

                            _order = {b:i for i,b in enumerate(BREVET_CHOICES)}


            

                            _part_df["__o"] = _part_df["Brevet"].map(_order)


            

                            _part_df = _part_df.sort_values(["__o","Achternaam","Voornaam"], na_position="last").drop(columns="__o")


            

                        except Exception:


            

                            _part_df = _part_df.sort_values(["Brevet","Achternaam","Voornaam"], na_position="last")


            

                


            

                    # Schrijf naar Excel (bovenbalk + lijst)


            

                    _out = io.BytesIO()


            

                    with pd.ExcelWriter(_out, engine="openpyxl") as _w:


            

                        _meta_df.to_excel(_w, index=False, sheet_name="Activiteit")


            

                        _part_df.to_excel(_w, index=False, sheet_name="Activiteit", startrow=3)


            

                        # Kolombreedtes instellen zoals voorbeeldbestand
                        try:
                            _ws = _w.book["Activiteit"] if "Activiteit" in _w.book.sheetnames else _w.book.active
                            _widths = [("A", None), ("B", None), ("C", None), ("D", 10.109375), ("E", 25.44140625), ("F", 41.77734375)]
                            for _col, _wpx in _widths:
                                if _wpx is not None:
                                    _ws.column_dimensions[_col].width = float(_wpx)
                        except Exception:
                            pass
                        _fname = f"activiteit_{(_d.strftime('%Y%m%d') if pd.notna(_d) else 'export')}.xlsx"


            

                    st.download_button("⬇️ Export naar Excel", data=_out.getvalue(), file_name=_fname, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=f"exp_{row['id']}")


            

                except Exception as __ex:


            

                    st.caption(f"Export niet mogelijk: {__ex}")
    if current_role() == "admin":    st.divider()
    # ==== Activiteiten bewerken / verwijderen (stabiele editor met formulier) ====
    df_del = activiteiten_list_df(upcoming=True).sort_values(["datum", "tijd"], na_position="last")
    if df_del.empty:
        st.caption("Geen toekomstige activiteiten.")
    else:
        # Bouw selectie-opties
        options, id_map = [], {}
        for _, r in df_del.iterrows():
            datum_str = pd.to_datetime(r["datum"]).strftime("%d/%m/%Y")
            tijd_str = f" · {r['tijd']}" if r.get("tijd") else ""
            loc_str = f" · {r['locatie']}" if r.get("locatie") else ""
            label = f"{datum_str}{tijd_str} · {r['titel']}{loc_str}"
            options.append(label); id_map[label] = r["id"]

        # Titelbalk met actieknoppen (geen radio/rode bol)
        t1, t2, t3 = st.columns([6, 1.2, 1.4])
        with t1:
            st.subheader("Activiteiten bewerken / verwijderen")
        with t2:
            clicked_edit = st.button("bewerken", key="act_hdr_edit")
        with t3:
            clicked_del = st.button("verwijderen", key="act_hdr_del")

        # Eerst activiteit kiezen (start leeg)
        _sel_placeholder = "— kies —"
        _edit_choice = st.selectbox("Kies activiteit", [_sel_placeholder] + options, index=0, key="mg_edit_select")
        _act_id = id_map.get(_edit_choice) if _edit_choice and _edit_choice != _sel_placeholder else None

        # Bewerkmodus toggelen
        if clicked_edit:
            if not _act_id:
                st.warning("Kies eerst een activiteit om te bewerken.")
            else:
                st.session_state["mg_edit_mode"] = True
                st.session_state["mg_edit_id"] = _act_id
                st.rerun()

        # Verwijderen (één item)
        if clicked_del:
            if not _act_id:
                st.warning("Kies eerst een activiteit om te verwijderen.")
            else:
                try:
                    run_db(lambda c: c.table("activiteiten").delete().eq("id", _act_id).execute(), what="activiteiten delete one")
                    st.success("Activiteit verwijderd.")
                    # reset
                    st.session_state.pop("mg_edit_mode", None)
                    st.session_state.pop("mg_edit_id", None)
                    st.session_state.pop("mg_edit_select", None)
                    st.rerun()
                except Exception as e:
                    st.error(f"Verwijderen mislukt: {e}")

        # Als selectie wijzigt tijdens bewerken, sluit editor
        if st.session_state.get("mg_edit_mode") and st.session_state.get("mg_edit_id") != _act_id:
            st.session_state.pop("mg_edit_mode", None)
            st.session_state.pop("mg_edit_id", None)

        # Editor blijft open met formulier
        if st.session_state.get("mg_edit_mode") and st.session_state.get("mg_edit_id") == _act_id:
            # haal rij
            _row = None
            try:
                _row = df_del.loc[df_del["id"] == _act_id].iloc[0].to_dict()
            except Exception:
                _row = None
            if _row is None:
                try:
                    _res = run_db(lambda c: c.table("activiteiten").select("*").eq("id", _act_id).limit(1).execute(), what="activiteiten fetch one")
                    _row = (_res.data or [])[0] if getattr(_res, "data", None) else {}
                except Exception:
                    _row = {}

            with st.form(key=f"adm_edit_form_{_act_id}"):
                ec1, ec2 = st.columns([2,1])
                with ec1:
                    etitel = st.text_input("Titel*", value=str(_row.get("titel") or ""), key=f"adm_e_title_{_act_id}")
                    eomschr = st.text_area("Omschrijving", value=str(_row.get("omschrijving") or ""), key=f"adm_e_om_{_act_id}")
                    eopm = st.text_area("Opmerkingen", value=str(_row.get("opmerkingen") or ""), key=f"adm_e_opm_{_act_id}")
                with ec2:
                    try:
                        _d = pd.to_datetime(_row.get("datum")).date() if _row.get("datum") else datetime.date.today()
                    except Exception:
                        _d = datetime.date.today()
                    try:
                        _t = pd.to_datetime(_row.get("tijd")).time() if _row.get("tijd") else None
                    except Exception:
                        _t = None
                    edatum = st.date_input("Datum*", value=_d, format="DD/MM/YYYY", key=f"adm_e_date_{_act_id}")
                    etijd = st.time_input("Tijd (optioneel)", value=_t, key=f"adm_e_time_{_act_id}")
                    _pl = plaatsen_list()
                    _loc = _row.get("locatie") or ""
                    _ix = 1 + _pl.index(_loc) if _loc in _pl else 0
                    eloc = st.selectbox("Locatie", ["— kies —"] + _pl, index=_ix, key=f"adm_e_loc_{_act_id}")

                # Organisator
                try:
                    _ld2 = leden_list_df()
                    _ld2 = _ld2.dropna(subset=['username']).copy() if isinstance(_ld2, pd.DataFrame) else pd.DataFrame([])
                except Exception:
                    _ld2 = pd.DataFrame([])
                _org_labels2, _org_map2 = ['— kies —'], {'— kies —': None}
                _cur_org = _row.get('organisator') or None
                _def_ix = 0
                if not _ld2.empty:
                    _ld2 = _ld2.sort_values(['voornaam','achternaam'])
                    for _ix2, _r2 in _ld2.iterrows():
                        _u2 = str(_r2.get('username') or '').strip()
                        if not _u2: continue
                        _lbl2 = f"{_r2.get('voornaam','').strip()} {_r2.get('achternaam','').strip()} ({_u2})".strip()
                        _org_labels2.append(_lbl2); _org_map2[_lbl2] = _u2
                        if _u2 == _cur_org: _def_ix = len(_org_labels2) - 1
                organisator_sel2 = st.selectbox('Organisator', _org_labels2, index=_def_ix, key=f"adm_e_org_{_act_id}")
                organisator_user2 = _org_map2.get(organisator_sel2)

                c1, c2 = st.columns([1,1])
                submit = c1.form_submit_button("Opslaan wijzigingen")
                cancel = c2.form_submit_button("Annuleren")

                if submit:
                    try:
                        payload = {
                            "titel": etitel.strip(),
                            "omschrijving": (eomschr or "").strip(),
                            "datum": edatum.isoformat(),
                            "tijd": (etijd.isoformat() if etijd else None),
                            "locatie": (None if not eloc or eloc == "— kies —" else eloc.strip()),
                            "opmerkingen": (eopm or "").strip() or None,
                            "organisator": (organisator_user2 or None)
                        }
                        activiteit_update(_act_id, payload)
                        st.success("Activiteit bijgewerkt.")
                        # Reset naar lege velden/keuze
                        st.session_state.pop("mg_edit_mode", None)
                        st.session_state.pop("mg_edit_id", None)
                        st.session_state.pop("mg_edit_select", None)
                        for _k in list(st.session_state.keys()):
                            if _k.startswith('adm_e_'):
                                del st.session_state[_k]
                        st.rerun()
                    except Exception as ex:
                        st.error(f"Opslaan mislukt: {ex}")
                if cancel:
                    st.session_state.pop("mg_edit_mode", None)
                    st.session_state.pop("mg_edit_id", None)
                    st.session_state.pop("mg_edit_select", None)
                    for _k in list(st.session_state.keys()):
                        if _k.startswith('adm_e_'):
                            del st.session_state[_k]
                    st.rerun()


    st.divider()
    col_w_title, col_w_toggle = st.columns([0.75, 0.25])
    with col_w_title:
        st.subheader("Wekelijkse mail — eerstvolgende 4 activiteiten")
    with col_w_toggle:
        if current_role() == "admin":
            try:
                _wk_enabled = settings_get_bool("weekly_mail_enabled", False)
            except Exception:
                _wk_enabled = False
            try:
                _wk_new = st.toggle("Automatisch versturen", value=_wk_enabled, key="wkmail_enabled")
            except Exception:
                _wk_new = st.checkbox("Automatisch versturen", value=_wk_enabled, key="wkmail_enabled_ck")
            if _wk_new != _wk_enabled:
                try:
                    settings_set_bool("weekly_mail_enabled", _wk_new)
                    st.toast("Weekmail ingeschakeld" if _wk_new else "Weekmail uitgeschakeld")
                    st.rerun()
                except Exception as _ex:
                    st.error(f"Opslaan mislukt: {_ex}")
        st.markdown('')
        col_wk1, col_wk2 = st.columns(2)
        with col_wk1:
            if st.button('Testmail naar mezelf', key='wkmail_test_btn'):
                ok, msg = weekmail_test_send('d.verbraeken@gmail.com')
                (st.success(msg) if ok else st.error(msg))
        with col_wk2:
            if st.button('Weekmail nu versturen', key='wkmail_send_btn'):
                ok, msg = weekmail_send_now()
                (st.success(msg) if ok else st.error(msg))

    st.caption("Gebruik als preview/export. Voor automatisch versturen op maandag 08:00 heb je een scheduler nodig.")
    df2 = activiteiten_list_df(upcoming=True).sort_values(["datum", "tijd"], na_position="last").head(4)
    if df2.empty: st.info("Geen komende activiteiten.")
    else:
        view = df2[["titel", "datum", "tijd", "locatie", "omschrijving", "opmerkingen"]].copy()
        view["datum"] = pd.to_datetime(view["datum"]).dt.strftime("%d/%m/%Y")
        st.dataframe(view, use_container_width=True, hide_index=True)
        out = io.BytesIO(); view.to_csv(out, index=False)
        st.download_button("⬇️ Exporteer CSV (mailbijlage)", data=out.getvalue(), file_name="weekmail_activiteiten.csv", mime="text/csv", key="weekmail_csv")

def page_duiken():
    require_role("admin", "user")
    if is_readonly(): st.warning("Read-only modus actief — opslaan uitgeschakeld.")
    appbar("duiken"); st.header("Duiken invoeren")
    # Reset 'Duiken invoeren' velden na succesvol bewaren
    if st.session_state.get("_reset_duiken", False):
        for _k in ["duiken_datum","duiken_duikcode","duiken_plaats","duiken_duikers","duiken_from_act"]:
            try:
                st.session_state.pop(_k, None)
            except Exception:
                pass
        st.session_state["_reset_duiken"] = False

    labels = duikers_labels()
    if not labels:
        try:
            _leden_df = run_db(lambda c: c.table("leden").select("voornaam, achternaam, duikbrevet, actief").execute(), what="leden voor duikers")
            _leden_df = pd.DataFrame(_leden_df.data or []); _labels_tmp = []
            if not _leden_df.empty:
                _f = _leden_df.copy(); _f["akt"] = _f.get("actief", True).fillna(True); _f["brev"] = _f.get("duikbrevet").fillna("")
                _f = _f[(_f["akt"] == True) & (_f["brev"] != "") & (_f["brev"] != "(geen)")]
                for _, _r in _f.iterrows():
                    _vn = (_r.get("voornaam") or "").strip(); _an = (_r.get("achternaam") or "").strip()
                    _lab = f"{_an}, {_vn}".strip(", "); 
                    if _lab: _labels_tmp.append(_lab)
                labels = sorted(set(_labels_tmp))
        except Exception: labels = []
    plaatsen = plaatsen_list()
    st.subheader("Overnemen uit activiteit")
    try: _acts = activiteiten_list_df(upcoming=False)
    except Exception: _acts = pd.DataFrame([])
    _act_opts = []; _act_map = {}
    if not _acts.empty:
        _acts_sorted = _acts.sort_values(["datum","tijd"], ascending=[False, True], na_position="last")
        for _, _r in _acts_sorted.iterrows():
            try: _d = pd.to_datetime(_r["datum"]).strftime('%d/%m/%Y')
            except Exception: _d = str(_r.get("datum") or "")
            _loc = (_r.get("locatie") or "").strip()
            _label = f"{_d} · {_r.get('titel') or ''}" + (f" · {_loc}" if _loc else "")
            _act_opts.append(_label); _act_map[_label] = _r.to_dict()
    _sel_act = st.selectbox("Overnemen uit activiteit", ["— kies activiteit —"] + _act_opts, index=0, key="duiken_from_act")
    _prefill_place = None; _prefill_duikers = []; _prefill_date = None
    if _sel_act != "— kies activiteit —":
        _ar = _act_map.get(_sel_act, {})
        _prefill_place = (_ar.get("locatie") or "").strip() or None
        try: _prefill_date = pd.to_datetime(_ar.get("datum")).date() if _ar.get("datum") else None
        except Exception: _prefill_date = None
        try: _s = signups_get(_ar["id"])
        except Exception: _s = pd.DataFrame([])
        if not _s.empty:
            _s_yes = _s[_s["status"] == "yes"]
            if not _s_yes.empty:
                _labels_set = set(labels); _pref = []
                for __, _sr in _s_yes.iterrows():
                    _vn = _an = ""; _un = _sr.get("username"); _lid = _sr.get("lid_id")
                    if _un:
                        _mem = leden_get_by_username(_un)
                        if _mem:
                            _vn = (_mem.get("voornaam") or "").strip(); _an = (_mem.get("achternaam") or "").strip()
                    if (not _vn and not _an) and _lid:
                        try:
                            _res = run_db(lambda c: c.table("leden").select("voornaam, achternaam").eq("id", _lid).limit(1).execute(), what="leden by id")
                            _mm = _res.data or []
                            if _mm:
                                _vn = (_mm[0].get("voornaam") or "").strip(); _an = (_mm[0].get("achternaam") or "").strip()
                        except Exception: pass
                    if _vn or _an:
                        _lab = f"{_an}, {_vn}".strip(", ")
                        if _lab in _labels_set: _pref.append(_lab)
                _prefill_duikers = sorted(set(_pref))
    datum_default = _prefill_date or datetime.date.today()
    datum = st.date_input("Datum", datum_default, format="DD/MM/YYYY", key="duiken_datum")
    duikcode = st.text_input("Duikcode (optioneel)", key="duiken_duikcode")
    _place_options = ["— kies —"] + plaatsen; _place_index = 0
    if _prefill_place and _prefill_place in plaatsen: _place_index = 1 + plaatsen.index(_prefill_place)
    plaats = st.selectbox("Duikplaats", _place_options, index=_place_index, key="duiken_plaats")
    sel_duikers = st.multiselect("Duikers", labels, default=_prefill_duikers, key="duiken_duikers")
    if current_role() == "admin":
        with st.expander("➕ Duikplaats toevoegen"):
            np2 = st.text_input("Nieuwe duikplaats", key="np_duiken")
            if st.button("Toevoegen", key="add_place_duiken"):
                if np2 and np2 not in plaatsen:
                    try: plaats_add(np2); st.success("Duikplaats toegevoegd."); st.rerun()
                    except Exception as e: st.error(f"Mislukt: {e}")
                else: st.warning("Leeg of al bestaand.")
    if st.button("Opslaan duik(en)", type="primary", disabled=(not sel_duikers or plaats == "— kies —" or is_readonly())):
        rows = [{"datum": datum.isoformat(), "plaats": plaats, "duiker": label.replace(", ", " "), "duikcode": duikcode or ""} for label in sel_duikers]
        try:

            duiken_insert(rows)

            st.success(f"{len(rows)} duik(en) opgeslagen.")

            st.session_state["_reset_duiken"] = True

            st.rerun()

        except Exception as e:

            st.error(f"Opslaan mislukt: {e}")

def page_overzicht():
    require_role("admin","user"); appbar("overzicht"); st.header("Overzicht duiken")
    df = duiken_fetch_df()
    if df.empty: st.info("Nog geen duiken."); return
    allowed_duiker_names = set([l.replace(", ", " ") for l in duikers_labels()])
    df = df[df["duiker"].isin(allowed_duiker_names)]
    if "id" not in df.columns: st.warning("Kolom 'id' ontbreekt in 'duiken' — verwijderen werkt niet."); df["id"] = None
    df["Datum"] = pd.to_datetime(df["datum"]).dt.date; df["Plaats"] = df["plaats"]; df["Duiker"] = df["duiker"]; df["Duikcode"] = df["duikcode"].fillna("")
    c1, c2, c3, c4 = st.columns([1,1,1,2])
    min_d, max_d = df["Datum"].min(), df["Datum"].max()
    try:
        import pandas as _pd
        from datetime import date as _date
        _min = _pd.to_datetime(min_d, dayfirst=True, errors='coerce'); _max = _pd.to_datetime(max_d, dayfirst=True, errors='coerce')
        if _pd.isna(_min) or _pd.isna(_max): min_d = max_d = _date.today()
        else: min_d = _min.date() if hasattr(_min, 'date') else _date.today(); max_d = _max.date() if hasattr(_max, 'date') else _date.today()
    except Exception: pass
    start_d = c1.date_input("Van datum", min_d, key="overz_start", format="DD/MM/YYYY")
    end_d = c2.date_input("Tot datum", max_d, key="overz_end", format="DD/MM/YYYY")
    pf = c2.selectbox("Duikplaats", ["Alle"] + sorted(df["Plaats"].dropna().unique().tolist()), index=0)
    cf = c3.selectbox("Duikcode", ["Alle"] + sorted([c if c else "—" for c in df["Duikcode"].unique().tolist()]), index=0)
    duikers = ["Alle"] + sorted(df["Duiker"].dropna().unique().tolist()); dfilt = c4.selectbox("Duiker", duikers, index=0)
    start, end = start_d, end_d
    f = df[(df["Datum"] >= start) & (df["Datum"] <= end)].copy()
    if pf != "Alle": f = f[f["Plaats"] == pf]
    if cf != "Alle": f = f[f["Duikcode"].replace({"": "—"}) == cf]
    if dfilt != "Alle": f = f[f["Duiker"] == dfilt]
    f = f.sort_values(["Datum","Plaats","Duikcode","Duiker","id"]).reset_index(drop=True)
    view = f[["Datum","Plaats","Duiker","Duikcode"]].copy(); view["Datum"] = pd.to_datetime(view["Datum"]).dt.strftime("%d/%m/%Y")
    st.dataframe(view, use_container_width=True, hide_index=True)
    st.divider(); st.subheader("Duiken verwijderen (huidige filter)")
    options, id_map = [], {}; f2 = f.copy(); f2["Datum"] = pd.to_datetime(f2["Datum"]).dt.date
    for _, r in f2.iterrows():
        dc = r.get("Duikcode") or r.get("duikcode") or ""; dc = dc if dc else "—"
        label = f"{r['Datum'].strftime('%d/%m/%Y')} · {r['Plaats']} · {r['Duiker']} · {dc}"
        lbl = label if label not in id_map else f"{label}  (#ID:{r['id']})"
        options.append(lbl); id_map[lbl] = r["id"]
    sel = st.multiselect("Selecteer te verwijderen duiken", options)
    if st.button("Verwijder geselecteerde", disabled=(len(sel)==0)):
        ids = [id_map[x] for x in sel if id_map[x] is not None]
        if not ids: st.warning("Geen geldige ID's.")
        else:
            try: duiken_delete_by_ids(ids); st.success(f"Verwijderd: {len(ids)} duik(en)."); st.rerun()
            except Exception as e: st.error(f"Verwijderen mislukt: {e}")
    st.divider(); st.subheader("Export (Excel)")
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as w: view.to_excel(w, index=False, sheet_name="Duiken")
    st.download_button("⬇️ Download Excel", data=out.getvalue(), file_name="duiken_export.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

def page_afrekening():
    require_role("admin","user","viewer"); appbar("afrekening"); st.header("Afrekening")
    df = duiken_fetch_df()
    if df.empty: st.info("Nog geen duiken."); return
    allowed_duiker_names = set([l.replace(", ", " ") for l in duikers_labels()])
    df = df[df["duiker"].isin(allowed_duiker_names)]
    if df.empty: st.info("Er zijn nog geen duiken voor geregistreerde duikers."); return
    df["Datum"] = pd.to_datetime(df["datum"]).dt.date; df["Plaats"] = df["plaats"]; df["Duiker"] = df["duiker"]
    c1, c2, c3, c4 = st.columns(4)
    min_d, max_d = df["Datum"].min(), df["Datum"].max()
    rng = c1.date_input("Periode", (min_d, max_d))
    bedrag = c2.number_input("Bedrag per duik (€)", min_value=0.0, step=0.5, value=5.0)
    pf = c3.selectbox("Duikplaats (optioneel)", ["Alle"] + sorted(df["Plaats"].dropna().unique().tolist()), index=0)
    blok = c4.number_input("Blokgrootte (€)", min_value=0.0, step=10.0, value=30.0)
    start, end = rng if isinstance(rng, tuple) else (min_d, max_d)
    m = (df["Datum"] >= start) & (df["Datum"] <= end)
    if pf != "Alle": m &= df["Plaats"] == pf
    s = df.loc[m].copy()
    if s.empty: st.warning("Geen duiken in de gekozen periode/filters."); return
    per = s.groupby("Duiker").size().reset_index(name="AantalDuiken")
    per["Bruto"] = (per["AantalDuiken"] * bedrag).round(2)
    try:
        ddf = run_db(lambda c: c.table("duikers").select("voornaam, achternaam, naam, rest_saldo").execute(), what="duikers join").data or []
        ddf = pd.DataFrame(ddf)
    except Exception: ddf = pd.DataFrame([])
    vns, ans, rests = [], [], []
    for disp in per["Duiker"].astype(str).tolist():
        vn, an, rest = "", "", 0.0
        if not ddf.empty:
            row = ddf.loc[ddf["naam"] == disp]
            if not row.empty:
                vn = (row.iloc[0].get("voornaam") or "").strip(); an = (row.iloc[0].get("achternaam") or "").strip(); rest = float(row.iloc[0].get("rest_saldo") or 0)
            else:
                parts = disp.split()
                if parts:
                    vn = parts[-1]
                    an = " ".join(parts[:-1])
                row2 = ddf.loc[(ddf["voornaam"].fillna("").str.strip()==vn) & (ddf["achternaam"].fillna("").str.strip()==an)]
                if not row2.empty: rest = float(row2.iloc[0].get("rest_saldo") or 0)
        else:
            parts = disp.split()
            if parts:
                vn = parts[-1]
                an = " ".join(parts[:-1])
        vns.append(vn); ans.append(an); rests.append(round(float(rest),2))
    per["Voornaam"]=vns; per["Achternaam"]=ans; per["RestOud"]=rests; per["Totaal"]=(per["Bruto"]+per["RestOud"]).round(2)
    def calc_blokken(total: float):
        if blok <= 0: return 0, 0.0, round(total, 2)
        n = math.floor(total / blok); uit = round(n * blok, 2); rest = round(total - uit, 2); return n, uit, rest
    rows=[]; 
    for _, r in per.iterrows():
        n, uit, rest = calc_blokken(float(r["Totaal"])); rows.append({**r.to_dict(), "Blokken": n, "UitTeBetalen": uit, "RestNieuw": rest})
    per = pd.DataFrame(rows).sort_values(["Achternaam","Voornaam","Duiker"], na_position="last").reset_index(drop=True)
    st.subheader("Afrekening per duiker")
    show_cols = ["Achternaam","Voornaam","AantalDuiken","Bruto","RestOud","Totaal","Blokken","UitTeBetalen","RestNieuw"]
    st.dataframe(per[show_cols], use_container_width=True, hide_index=True)

    st.divider(); st.subheader("Export (Excel)")
    _out = io.BytesIO()
    _export = per[show_cols].copy()
    try:
        with pd.ExcelWriter(_out, engine="openpyxl") as _w:
            # Meta bovenaan met periode en parameters
            _meta = pd.DataFrame([{
                "Periode start": (pd.to_datetime(start).strftime("%d/%m/%Y") if pd.notna(pd.to_datetime(start, errors="coerce")) else str(start)),
                "Periode einde": (pd.to_datetime(end).strftime("%d/%m/%Y") if pd.notna(pd.to_datetime(end, errors="coerce")) else str(end)),
                "Bedrag per duik (€)": float(bedrag),
                "Blokgrootte (€)": float(blok)
            }])
            _meta.to_excel(_w, index=False, sheet_name="Afrekening")
            _export.to_excel(_w, index=False, sheet_name="Afrekening", startrow=3)
            # Kolombreedtes voor leesbaarheid
            try:
                _ws = _w.book["Afrekening"] if "Afrekening" in _w.book.sheetnames else _w.book.active
                _ws.column_dimensions["A"].width = 18
                _ws.column_dimensions["B"].width = 18
                _ws.column_dimensions["C"].width = 14
                _ws.column_dimensions["D"].width = 12
                _ws.column_dimensions["E"].width = 12
                _ws.column_dimensions["F"].width = 12
                _ws.column_dimensions["G"].width = 14
                _ws.column_dimensions["H"].width = 14
                _ws.column_dimensions["I"].width = 12
            except Exception:
                pass
        _fname = f"afrekening_{(pd.to_datetime(start).strftime('%Y%m%d') if pd.notna(pd.to_datetime(start, errors='coerce')) else 'start')}_{(pd.to_datetime(end).strftime('%Y%m%d') if pd.notna(pd.to_datetime(end, errors='coerce')) else 'einde')}.xlsx"
        st.download_button("⬇️ Exporteer naar Excel", data=_out.getvalue(), file_name=_fname, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="exp_afrekening")
    except Exception as __ex:
        st.caption(f"Export niet mogelijk: {__ex}")

    if current_role() in {"admin","user","viewer"}:
        st.divider(); st.subheader("Markeer als betaald / update restsaldo")
        per["select"]=False
        for i in range(len(per)):
            label=f"{per.at[i,'Achternaam']}, {per.at[i,'Voornaam']}"; per.at[i,"select"]=st.checkbox(label, key=f"sel_pay_{i}")
        if st.button("Markeer geselecteerde als betaald"):
            try:
                sel = per[per["select"]==True]
                if sel.empty: st.warning("Geen duikers geselecteerd.")
                else:
                    for _, r in sel.iterrows():
                        row = {"voornaam": (r["Voornaam"] or "").strip(), "achternaam": (r["Achternaam"] or "").strip(),
                               "periode_start": start, "periode_end": end, "bedrag_per_duik": float(bedrag), "blokgrootte": float(blok),
                               "aantal_duiken": int(r["AantalDuiken"]), "bruto_bedrag": float(r["Bruto"]), "rest_oud": float(r["RestOud"]),
                               "blokken": int(r["Blokken"]), "uit_te_betalen": float(r["UitTeBetalen"]), "rest_nieuw": float(r["RestNieuw"]),
                               "betaald_op": dt.utcnow().isoformat() }
                        afrekening_insert(row)
                        if row["voornaam"] or row["achternaam"]:
                            run_db(lambda c: c.table("duikers").update({"rest_saldo": float(r["RestNieuw"])}).eq("voornaam", row["voornaam"]).eq("achternaam", row["achternaam"]).execute(), what="duikers update rest_saldo")
                    st.success("Afrekening geregistreerd."); st.rerun()
            except Exception as e: st.error(f"Registratie mislukt: {e}")

def page_beheer():
    require_role("admin"); appbar("beheer"); st.header("Beheer")
    tabs = st.tabs(["Ledenbeheer", "Duikers", "Duikplaatsen", "Back-up/Export"])
    with tabs[0]: page_ledenbeheer()
    with tabs[1]:
        res = run_db(lambda c: c.table("duikers").select("voornaam, achternaam, naam, rest_saldo").execute(), what="duikers select (beheer)")
        ddf = pd.DataFrame(res.data or []); st.subheader("Duikers (afgeleid uit leden met duikbrevet)")
        if not ddf.empty:
            view = ddf.rename(columns={"voornaam": "Voornaam", "achternaam": "Achternaam", "rest_saldo": "Rest (start)"})
            st.dataframe(view, use_container_width=True, hide_index=True)
        else: st.caption("Nog geen duikers — geef duikbrevet aan een lid in Ledenbeheer.")
    with tabs[2]:
        st.subheader("Duikplaatsen"); pl = plaatsen_list()
        st.dataframe(pd.DataFrame({"Plaats": pl}), use_container_width=True, hide_index=True)
        np = st.text_input("Nieuwe duikplaats", key="np_beheer")
        if st.button("Toevoegen", key="add_place_beheer", disabled=is_readonly()):
            if np and np not in pl:
                try: plaats_add(np); st.success("Duikplaats toegevoegd."); st.rerun()
                except Exception as e: st.error(f"Mislukt: {e}")
            else: st.warning("Leeg of al bestaand.")
    with tabs[3]:
        st.subheader("Back-up (Excel)")
        if st.button("Maak back-up"):
            out = io.BytesIO()
            duikers = run_db(lambda c: c.table("duikers").select("*").execute(), what="duikers select (backup)")
            plaatsen_df = run_db(lambda c: c.table("duikplaatsen").select("*").execute(), what="duikplaatsen select (backup)")
            duiken = run_db(lambda c: c.table("duiken").select("*").execute(), what="duiken select (backup)")
            leden = run_db(lambda c: c.table("leden").select("*").execute(), what="leden select (backup)")
            df_duikers = pd.DataFrame(duikers.data or []); df_plaatsen = pd.DataFrame(plaatsen_df.data or [])
            df_duiken = pd.DataFrame(duiken.data or []); df_leden = pd.DataFrame(leden.data or [])
            stamp = dt.utcnow().strftime("%Y%m%d_%H%M%S")
            with pd.ExcelWriter(out, engine="openpyxl") as w:
                df_duikers.to_excel(w, index=False, sheet_name="duikers")
                df_plaatsen.to_excel(w, index=False, sheet_name="duikplaatsen")
                df_duiken.to_excel(w, index=False, sheet_name="duiken")
                df_leden.to_excel(w, index=False, sheet_name="leden")
            st.download_button("⬇️ Download back-up", data=out.getvalue(), file_name=f"anww_backup_{stamp}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

def main():
    try:
        if auth_count_admins() == 0 and "auth_user" not in st.session_state:
            page_setup_first_admin_username(); return
    except Exception:
        st.error("Tabel 'auth_local' ontbreekt. Voer de 1-klik SQL eerst uit."); st.stop()
    if "auth_user" not in st.session_state:
        page_login_username(); return
    role = current_role()
    if role == "admin":
        tabs = st.tabs(["Activiteiten","Duiken invoeren","Overzicht","Afrekening","Beheer","Mijn profiel"])
        with tabs[0]: page_activiteiten()
        with tabs[1]: page_duiken()
        with tabs[2]: page_overzicht()
        with tabs[3]: page_afrekening()
        with tabs[4]: page_beheer()
        with tabs[5]: page_profiel()
    elif role == "user":
        tabs = st.tabs(["Activiteiten","Duiken invoeren","Overzicht","Afrekening","Mijn profiel","Leden"])
        with tabs[0]: page_activiteiten()
        with tabs[1]: page_duiken()
        with tabs[2]: page_overzicht()
        with tabs[3]: page_afrekening()
        with tabs[4]: page_profiel()
        with tabs[5]: page_leden()
    elif role == "member":
        tabs = st.tabs(["Activiteiten","Mijn profiel","Leden"])
        with tabs[0]: page_activiteiten()
        with tabs[1]: page_profiel()
        with tabs[2]: page_leden()
    else:
        tabs = st.tabs(["Afrekening","Mijn profiel"])
        with tabs[0]: page_afrekening()
        with tabs[1]: page_profiel()
if __name__ == "__main__":
    main()

# --- Settings helpers (weekplanner toggle) ---
def settings_get_bool(key: str, default: bool = False) -> bool:
    try:
        res = run_db(lambda c: c.table("settings").select("value_bool").eq("key", key).maybe_single().execute(),
                     what="settings get")
        data = getattr(res, "data", None) or {}
        val = data.get("value_bool", None)
        return bool(val) if val is not None else default
    except Exception:
        return default

def settings_set_bool(key: str, value: bool) -> None:
    if is_readonly():
        return
    run_db(lambda c: c.table("settings").upsert({"key": key, "value_bool": bool(value)}).execute(),
           what="settings set")


def weekmail_test_send(to_email: str) -> tuple[bool, str]:
    """Call Supabase Edge Function 'weekmail' in dry-run mode to send only to 'to_email'."""
    # Haal Supabase URL en anon key uit secrets of environment
    supa = st.secrets.get("supabase", {}) if hasattr(st, "secrets") else {}
    base_url = supa.get("url") or os.getenv("SUPABASE_URL", "")
    anon_key = supa.get("anon_key") or os.getenv("SUPABASE_ANON_KEY", "")
    url = (base_url.rstrip('/') + "/functions/v1/weekmail") if base_url else ""
    if not url or not anon_key:
        return False, "SUPABASE_URL/ANON_KEY ontbreekt in secrets/env."
    payload = {"dryRun": True, "to": to_email}
    if APP_PUBLIC_URL:
        payload["appUrl"] = APP_PUBLIC_URL
    try:
        resp = requests.post(
            url,
            headers={
                "Authorization": f"Bearer {anon_key}",
                "Content-Type": "application/json",
            },
            json=payload,
            timeout=30,
        )
        if resp.status_code >= 400:
            return False, f"HTTP {resp.status_code}: {resp.text[:300]}"
        data = resp.json()
        if not data.get("ok", False):
            return False, f"Mislukt: {data}"
        return True, "Testmail verstuurd."
    except Exception as ex:
        return False, f"Fout bij aanroepen Edge Function: {ex}"


def weekmail_send_now() -> tuple[bool, str]:
    """Call Supabase Edge Function 'weekmail' om de echte weekmail te versturen."""
    supa = st.secrets.get("supabase", {}) if hasattr(st, "secrets") else {}
    base_url = supa.get("url") or os.getenv("SUPABASE_URL", "")
    anon_key = supa.get("anon_key") or os.getenv("SUPABASE_ANON_KEY", "")
    url = (base_url.rstrip('/') + "/functions/v1/weekmail") if base_url else ""
    if not url or not anon_key:
        return False, "SUPABASE_URL/ANON_KEY ontbreekt in secrets/env."
    payload = {"dryRun": False}
    if APP_PUBLIC_URL:
        payload["appUrl"] = APP_PUBLIC_URL
    try:
        resp = requests.post(
            url,
            headers={
                "Authorization": f"Bearer {anon_key}",
                "Content-Type": "application/json",
            },
            json=payload,
            timeout=60,
        )
        if resp.status_code >= 400:
            return False, f"HTTP {resp.status_code}: {resp.text[:300]}"
        data = resp.json()
        if not data.get("ok", False):
            return False, f"Mislukt: {data}"
        return True, "Weekmail verstuurd."
    except Exception as ex:
        return False, f"Fout bij aanroepen Edge Function: {ex}"
